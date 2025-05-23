import sqlite3
from jinja2 import Template
import yaml
import pandas as pd
from openpyxl import load_workbook
import os

# Load configuration from YAML
with open('config.yaml', 'r') as f:
    config = yaml.safe_load(f)

school_year_dash = config['years']
school_year_splat = [yr.replace('-', '_') for yr in school_year_dash]
file_path_prefix = config['file_path_prefix']
pa_school_output_file_name = f"{file_path_prefix}/{config['pa_school_output_file_name']}"
db_file = f"{file_path_prefix}/{config['db_file_name']}"

def write_dataframes(filepath, sheet_name, dataframes, header_row=1, start_row=2):
    """
    Unified writer for Excel (.xlsx) or CSV (.csv), with adjustable header and data start rows.
    Applies Excel-escaping to string values to prevent automatic formatting (e.g., 9-12 → date).
    """
    def escape_for_excel(val):
        if pd.isnull(val):
            return ''
        val_str = str(val)
        if '-' in val_str or '/' in val_str:
            return f'="{val_str}"'  # Excel will treat this as a string
        return val_str

    is_csv = filepath.lower().endswith(".csv")
    base_path, ext = os.path.splitext(filepath)
    actual_path = f"{base_path}.{sheet_name}{ext}" if is_csv else filepath
    use_all_columns = False
    header = []

    # === Header extraction ===
    if is_csv:
        try:
            with open(filepath, "r", encoding="utf-8") as f:
                for i, line in enumerate(f, start=1):
                    if i == header_row:
                        header = [h.strip() for h in line.strip().split(",") if h.strip()]
                        break
            if not header:
                print(f"Note: No header found in line {header_row} of '{filepath}'. Writing all DataFrame columns.")
                use_all_columns = True
        except FileNotFoundError:
            print(f"Warning: CSV file '{filepath}' not found. Writing all DataFrame columns.")
            use_all_columns = True
    else:
        try:
            book = load_workbook(filepath, read_only=True)
            if sheet_name not in book.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in workbook '{filepath}'.")
            sheet = book[sheet_name]
            header = [cell.value for cell in sheet[header_row] if cell.value]
            if not header:
                print(f"Note: No header found in row {header_row} of sheet '{sheet_name}'. Writing all DataFrame columns.")
                use_all_columns = True
        except FileNotFoundError:
            raise FileNotFoundError(f"Excel file '{filepath}' not found.")
        except Exception as e:
            raise RuntimeError(f"Failed to read header from Excel file '{filepath}': {e}")

    # === CSV Writer ===
    if is_csv:
        filtered_dfs = []

        for i, df in enumerate(dataframes):
            df_cols = list(df.columns)
            common_columns = df_cols if use_all_columns else [col for col in header if col in df_cols]
            skipped_columns = [col for col in df_cols if col not in common_columns]

            if skipped_columns:
                print(f"[DataFrame {i+1}] Skipped columns not in '{sheet_name}': {skipped_columns}")
            if not common_columns:
                print(f"[DataFrame {i+1}] No matching columns found. Skipping.")
                continue

            print(f"Writing columns to [DataFrame {i+1}] in CSV output '{actual_path}': {common_columns}")
            filtered_df = df[common_columns].copy()

            for col in common_columns:
                filtered_df[col] = filtered_df[col].apply(escape_for_excel)

            filtered_dfs.append(filtered_df)

        if filtered_dfs:
            combined_df = pd.concat(filtered_dfs, ignore_index=True)
            if not use_all_columns:
                combined_df = combined_df.reindex(columns=header)
            combined_df.to_csv(actual_path, index=False)
            print(f"Successfully wrote CSV for '{sheet_name}' to '{actual_path}'.")
        else:
            print("No data written: all dataframes were empty or had no matching columns.")

    else:
        # === Excel Writer ===
        book = load_workbook(filepath)
        sheet = book[sheet_name]

        if sheet.max_row > (start_row - 1):
            sheet.delete_rows(start_row, sheet.max_row - (start_row - 1))

        for i, df in enumerate(dataframes):
            df_cols = list(df.columns)
            common_columns = df_cols if use_all_columns else [col for col in header if col in df_cols]
            skipped_columns = [col for col in df_cols if col not in common_columns]

            if skipped_columns:
                print(f"[DataFrame {i+1}] Skipped columns not in '{sheet_name}': {skipped_columns}")
            if not common_columns:
                print(f"[DataFrame {i+1}] No matching columns found. Skipping.")
                continue

            print(f"Writing columns to [DataFrame {i+1}] in sheet '{sheet_name}': {common_columns}")
            col_map = {col: idx + 1 for idx, col in enumerate(common_columns)}

            for r_idx, row in df.iterrows():
                for col in common_columns:
                    val = row.get(col, '')
                    val = escape_for_excel(val)
                    cell = sheet.cell(row=start_row + r_idx, column=col_map[col], value=val)
                    cell.number_format = '@'  # force Excel to treat as text

        book.save(filepath)
        print(f"Successfully wrote Excel sheet '{sheet_name}' in '{filepath}'.")
        
# Prepare to store the results
offerings_dfs = []
school_list_dfs = []

# Connect to SQLite
conn = sqlite3.connect(db_file)

# SQL files and their corresponding output targets
scripts = {
    'pa_school_offerings_jinja.sql': offerings_dfs,
    'pa_school_list_jinja.sql': school_list_dfs
}

for script_name, target_list in scripts.items():
    with open(script_name, 'r') as file:
        template = Template(file.read())

    for i in range(len(school_year_dash)):
        dash = school_year_dash[i]
        splat = school_year_splat[i]

        rendered_sql = template.render(
            school_year_dash=dash,
            school_year_splat=splat
        )

        cursor = conn.cursor()
        try:
            cursor.execute(rendered_sql)
            results = cursor.fetchall()
            column_names = [desc[0] for desc in cursor.description]
            df = pd.DataFrame(results, columns=column_names)
            target_list.append(df)
        except sqlite3.Error as e:
            print(f"SQL error while executing {script_name} for {splat}: {e}")

conn.close()

# Write results to Excel
write_dataframes(pa_school_output_file_name, "Tab 1 - School Offerings", offerings_dfs, header_row=2, start_row=3)
write_dataframes(pa_school_output_file_name, "School list", school_list_dfs, header_row=2, start_row=3)

print("Done.")
