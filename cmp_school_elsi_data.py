import pandas as pd
import yaml # pip install pyyaml
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor
from rapidfuzz import process, fuzz # pip install rapidfuzz
import re

# Load configuration from YAML
with open('config.yaml', 'r') as f:
    config = yaml.safe_load(f)
    
cmp_file_path_prefix = config['cmp_file_path_prefix']
elsi_school_file_name = f"{cmp_file_path_prefix}/{config['elsi_school_file_name']}"
elsi_district_file_name = f"{cmp_file_path_prefix}/{config['elsi_district_file_name']}"
cmp_output_file = f"{cmp_file_path_prefix}/{config['cmp_output_file_name']}"

output_school_name_col = config['output_school_name_col']
output_district_name_col = config['output_district_name_col']
output_school_id_col = config['output_school_id_col']
output_district_id_col = config['output_district_id_col']
output_low_grade_band_col = config['output_low_grade_band_col']
output_high_grade_band_col = config['output_high_grade_band_col']

elsi_school_col = config['elsi_school_col']
elsi_district_col = config['elsi_district_col']
elsi_school_id_col = config['elsi_school_id_col']
elsi_district_id_col = config['elsi_district_id_col']
elsi_low_grade_band = config['elsi_low_grade_band']
elsi_high_grade_band = config['elsi_high_grade_band']

def normalize_name(name):
    if not name:
        return ''
    name = name.lower()
    name = re.sub(r'[^\w\s-]', '', name)  # remove punctuation
    name = re.sub(r'\s+', ' ', name)     # collapse whitespace
    return name.strip()

def get_id_from_name(name, source_df, name_col, id_col, threshold=70):
    if not name:
        return None, None

    # Normalize the query and choices
    normalized_query = normalize_name(name)
    choices = source_df[name_col].astype(str)
    normalized_choices = choices.map(normalize_name).tolist()

    # Use normalized names for matching
    best_match = process.extractOne(normalized_query, normalized_choices,
                                     scorer=fuzz.token_sort_ratio,
                                     score_cutoff=threshold)

    if best_match:
        match, score, match_index = best_match
        #print(f"Matched {name} to {match}")
        return source_df.iloc[match_index][id_col], match_index

    print(f"No match for: '{name}' (normalized: '{normalized_query}')")
    
    return None, None
    
def match_row(cells, idx, col_idx, school_df, district_df):
    school_name = cells[col_idx['school_name'] - 1].value if col_idx['school_name'] else None
    district_name = cells[col_idx['district_name'] - 1].value if col_idx['district_name'] else None

    school_id = school_idx = None
    if school_name:
        school_id, school_idx = get_id_from_name(school_name, school_df,
                                                 elsi_school_col,
                                                 elsi_school_id_col)

    district_id = None
    if district_name:
        district_id, _ = get_id_from_name(district_name, district_df,
                                          elsi_district_col,
                                          elsi_district_id_col)

    low_band = high_band = None
    if school_idx is not None:
        if col_idx.get('low_grade_band'):
            low_band = school_df.iloc[school_idx][elsi_low_grade_band]
        if col_idx.get('high_grade_band'):
            high_band = school_df.iloc[school_idx][elsi_high_grade_band]

    return {
        'cells': cells,
        'school_id_col': col_idx.get('school_id'), 'school_id': school_id,
        'district_id_col': col_idx.get('district_id'), 'district_id': district_id,
        'low_band_col': col_idx.get('low_grade_band'), 'low_band': low_band,
        'high_band_col': col_idx.get('high_grade_band'), 'high_band': high_band,
    }

# Load source Excel data (data starts on row 7 in the source files)
district_df = pd.read_excel(elsi_district_file_name, skiprows=6)
school_df = pd.read_excel(elsi_school_file_name, skiprows=6)

# Load output file with openpyxl
wb = load_workbook(cmp_output_file)

for output_sheet_name in ["School Pop. Data", "School CS Data"]:
    ws = wb[output_sheet_name]

    # Load header row from output file to find column indices
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    # Get column indices (1-based for openpyxl)
    column_map = {
        'school_name': output_school_name_col,
        'district_name': output_district_name_col,
        'school_id': output_school_id_col,
        'district_id': output_district_id_col,
        'low_grade_band': output_low_grade_band_col,
        'high_grade_band': output_high_grade_band_col,
    }

    # Get column indices only if present in header
    col_idx = {
        key: header.index(col_name) + 1 if col_name in header else None
        for key, col_name in column_map.items()
    }
    
    # Phase 1: Collect all data rows (value-only)
    rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row)) # Start from row 2 (since row 1 is the header)

    # Phase 2: Run matching in parallel
    results = []
    with ThreadPoolExecutor(max_workers=8) as executor:
        futures = [
            executor.submit(match_row, row, idx + 2, col_idx, school_df, district_df)
            for idx, row in enumerate(rows)
        ]
        for future in futures:
            results.append(future.result())

    # Phase 3: Write results to Excel (sequential)
    for result in results:
        cells = result['cells']
        if result['school_id_col'] and result['school_id'] is not None:
            cells[result['school_id_col'] - 1].value = result['school_id']
        if result['district_id_col'] and result['district_id'] is not None:
            cells[result['district_id_col'] - 1].value = result['district_id']
        if result['low_band_col'] and result['low_band'] is not None:
            cells[result['low_band_col'] - 1].value = result['low_band']
        if result['high_band_col'] and result['high_band'] is not None:
            cells[result['high_band_col'] - 1].value = result['high_band']

    # Save in-place 
    wb.save(cmp_output_file)
