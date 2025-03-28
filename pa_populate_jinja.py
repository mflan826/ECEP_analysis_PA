import sqlite3
from jinja2 import Template
import yaml
import pandas as pd
from openpyxl import load_workbook

# Load configuration from YAML
with open('config.yaml', 'r') as f:
    config = yaml.safe_load(f)

school_year_dash = config['years']
school_year_splat = [yr.replace('-', '_') for yr in school_year_dash]
file_path_prefix = config['file_path_prefix']
pa_output_file = f"{file_path_prefix}/{config['pa_output_file_name']}"
db_file = f"{file_path_prefix}/{config['db_file_name']}"

def clear_and_append_dataframes_to_excel(filepath, sheet_name, dataframes):
    try:
        book = load_workbook(filepath)
    except FileNotFoundError:
        raise FileNotFoundError(f"File '{filepath}' not found.")

    if sheet_name not in book.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in the workbook.")

    sheet = book[sheet_name]

    # Read existing header (row 2)
    header = [cell.value for cell in sheet[2] if cell.value]
    if not header:
        raise ValueError(f"No header found in row 2 of sheet '{sheet_name}'.")

    # Clear all rows except the header
    max_row = sheet.max_row
    if max_row > 2:
        sheet.delete_rows(3, max_row - 1)

    start_row = 3

    for i, df in enumerate(dataframes):
        df_cols = list(df.columns)
        common_columns = [col for col in header if col in df_cols]
        skipped_columns = [col for col in df_cols if col not in header]

        if skipped_columns:
            print(f"[DataFrame {i+1}] Skipped columns not in '{sheet_name}': {skipped_columns}")
        if not common_columns:
            print(f"[DataFrame {i+1}] No matching columns found for sheet '{sheet_name}'. Skipping.")
            continue
        else:
            print(f"Writing columns to [DataFrame {i+1}] in '{sheet_name}': {common_columns}")

        header_column_map = {col_name: idx + 1 for idx, col_name in enumerate(header)}

        for r_idx, row in df.iterrows():
            for col in common_columns:
                value = row.get(col, '')
                col_idx = header_column_map[col]
                sheet.cell(row=start_row + r_idx, column=col_idx, value=value)

    book.save(filepath)

# Prepare to store the results
offerings_dfs = []
school_list_dfs = []

# Connect to SQLite
conn = sqlite3.connect(db_file)

# SQL files and their corresponding output targets
scripts = {
    'pa_school_offerings_jinja.sql': offerings_dfs,
    'pa_school_list_jinja.sql': school_list_dfs
}

for script_name, target_list in scripts.items():
    with open(script_name, 'r') as file:
        template = Template(file.read())

    for i in range(len(school_year_dash)):
        dash = school_year_dash[i]
        splat = school_year_splat[i]

        rendered_sql = template.render(
            school_year_dash=dash,
            school_year_splat=splat
        )

        cursor = conn.cursor()
        try:
            cursor.execute(rendered_sql)
            results = cursor.fetchall()
            column_names = [desc[0] for desc in cursor.description]
            df = pd.DataFrame(results, columns=column_names)
            target_list.append(df)
        except sqlite3.Error as e:
            print(f"SQL error while executing {script_name} for {splat}: {e}")

conn.close()

# Write results to Excel
clear_and_append_dataframes_to_excel(pa_output_file, "Tab 1 - School Offerings", offerings_dfs)
clear_and_append_dataframes_to_excel(pa_output_file, "School list", school_list_dfs)

print("Done.")
