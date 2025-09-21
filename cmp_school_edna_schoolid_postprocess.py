import traceback
import pandas as pd
from openpyxl import load_workbook
from tqdm import tqdm
import re, time, html
from typing import Optional, Tuple, List
from urllib.parse import urljoin, quote_plus
import requests
from bs4 import BeautifulSoup, Tag
import os
import yaml

try:
    with open('config.yaml', 'r') as f:
        config = yaml.safe_load(f)
except Exception as e:
    print(f"[config] Failed to load config.yaml: {e}")
    traceback.print_exc()
    raise

# ==============================
# Configuration
# ==============================
file_path_prefix       = config['file_path_prefix']
CMP_FILENAME = f"{file_path_prefix}/CMP_Data_Populated.xlsx"
OUTPUT_FILENAME = f"{file_path_prefix}/CMP_Data_Populated - Updated.xlsx"
EDNA_CACHE_CSV = f"{file_path_prefix}/{config['edna_cache']}"
CMP_SHEETS = ["School Pop. Data", "School CS Data"]
ENABLE_FUZZY_MATCH = False   # set to False to disable fuzzy matching

# ==============================
# Exact-match only by default; fuzzy fallback
# ==============================
try:
    # RapidFuzz is faster and actively maintained
    from rapidfuzz import process as _rf_process
    _USE_RAPIDFUZZ = True
except Exception:
    _USE_RAPIDFUZZ = False
    try:
        from fuzzywuzzy import process as _fw_process
    except Exception:
        _fw_process = None  # If absent, we'll detect and behave accordingly

FUZZY_THRESHOLD = 60

# ==============================
# Utilities
# ==============================
def _force_screens_url(url_or_href: str) -> str:
    """
    If a URL (or href) points to wfInstitutionDetails.aspx without the /Screens/ prefix,
    normalize it to '/Screens/wfInstitutionDetails.aspx?...'. Works for:
      - full URLs: http://www.edna.pa.gov/wfInstitutionDetails.aspx?ID=12345
      - paths:     /wfInstitutionDetails.aspx?ID=12345
      - bare:      wfInstitutionDetails.aspx?ID=12345
    Leaves other pages (e.g., wfSchools.aspx) unchanged.
    """
    s = (url_or_href or "").strip()
    if not s:
        return s
    low = s.lower()
    if "wfinstitutiondetails.aspx" not in low:
        return s

    # Strip domain, keep path+query if a full URL
    try:
        from urllib.parse import urlparse
        p = urlparse(s)
        if p.scheme and p.netloc:
            s = p.path + (("?" + p.query) if p.query else "")
    except Exception:
        pass

    # Ensure leading slash
    if not s.startswith("/"):
        s = "/" + s

    # Ensure '/Screens/' prefix
    if not s.lower().startswith("/screens/"):
        s = "/Screens" + s

    # Collapse accidental repeats or double slashes
    s = s.replace("//", "/").replace("/Screens/Screens/", "/Screens/")
    # Reattach host
    return urljoin(EDNA_BASE, s)

def _normalize_detail_url(href_or_url: str) -> str:
    """
    EDNA sometimes emits links like '/wfInstitutionDetails.aspx?ID=46960'
    but the working endpoint is '/Screens/wfInstitutionDetails.aspx?ID=46960'.
    Normalize any variant to the canonical '/Screens/...' path.
    """
    try:
        s = (href_or_url or "").strip()
        if not s:
            return s

        # Only touch paths that point at wfInstitutionDetails.aspx
        # Handle cases with/without leading slash, with querystring, etc.
        # Examples we fix:
        #   'wfInstitutionDetails.aspx?ID=46960'
        #   '/wfInstitutionDetails.aspx?ID=46960'
        #   'Screens/wfInstitutionDetails.aspx?ID=46960'  (add leading slash)
        # Leave others alone (e.g., wfSchools.aspx, search pages, etc.)
        lower = s.lower()
        if "wfinstitutiondetails.aspx" in lower:
            # strip leading domain if someone passed a full URL
            # then operate on just the path+query
            try:
                from urllib.parse import urlparse
                p = urlparse(s)
                if p.scheme and p.netloc:
                    s = p.path + (("?" + p.query) if p.query else "")
            except Exception:
                pass

            # ensure leading slash
            if not s.startswith("/"):
                s = "/" + s

            # ensure '/Screens/' prefix
            if not s.lower().startswith("/screens/"):
                # replace leading '/' with '/Screens/'
                # but if it already starts with '/Screens/' this is no-op
                s = "/Screens" + s

            # collapse any accidental double '/Screens/Screens/...'
            s = s.replace("//", "/")
            s = s.replace("/Screens/Screens/", "/Screens/")

        return s
    except Exception:
        traceback.print_exc()
        return href_or_url

def norm(s: str) -> str:
    return s.strip() if isinstance(s, str) else ""

def _status_norm(s: str) -> str:
    return re.sub(r"\s+", " ", str(s or "")).strip().lower()

def _status_canonicalize(s: str) -> str:
    """
    Map various forms to 'Open' or 'Closed' when obvious.
    Falls back to the original string if unknown.
    """
    sn = _status_norm(s)
    if sn.startswith("open"):
        return "Open"
    if sn.startswith("closed"):
        return "Closed"
    return s or ""

def ensure_headers(df: pd.DataFrame, required: list, ctx: str):
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"[{ctx}] Missing required columns: {missing}. Found: {list(df.columns)}")

def build_header_map(ws):
    """Return dict: header_text -> 1-based column index, from the first row in a worksheet."""
    header_cells = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    return {str(v): i + 1 for i, v in enumerate(header_cells) if v is not None}

def _extract_one(query: str, choices: list[str]):
    """
    Backend-agnostic extractOne(query, choices) -> (best_match, score)
    Uses RapidFuzz if available; otherwise fuzzywuzzy; otherwise returns (None, 0).
    Scores are on 0..100 scale in both libs.
    """
    if _USE_RAPIDFUZZ:
        # RapidFuzz returns (choice, score, index)
        res = _rf_process.extractOne(query, choices)
        if res:
            return res[0], float(res[1])
        return None, 0.0
    elif _fw_process is not None:
        res = _fw_process.extractOne(query, choices)
        if res:
            return res[0], float(res[1])
        return None, 0.0
    else:
        return None, 0.0

# ==============================
# Edna Lookup
# ==============================

EDNA_BASE = "http://www.edna.pa.gov"
CURRENTNAME_SEARCH_TEMPLATE = (
    "http://www.edna.pa.gov/Screens/wfSearchEntityResults.aspx?"
    "AUN=&SchoolBranch=&CurrentName={CURRENT}&City=&HistoricalName=&IU=-1&CID=-1&"
    "CategoryIDs=3%2c4%2c2%2c1%2c6%2c7%2c29%2c27%2c28%2c55%2c30%2c31%2c14%2c11%2c12%2c9%2c17%2c15%2c18%2c"
    "46%2c47%2c40%2c34%2c56%2c45%2c36%2c44%2c35%2c38%2c59%2c999%2c32%2c33%2c37%2c49%2c57%2c52%2c22%2c20%2c19%2c58%2c53%2c"
    "3%2c4%2c2%2c1%2c6%2c7%2c29%2c27%2c28%2c55%2c30%2c31%2c14%2c11%2c12%2c9%2c17%2c15%2c18%2c46%2c47%2c40%2c34%2c56%2c45%2c36%2c44%2c35%2c38%2c59%2c999%2c32%2c33%2c37%2c49%2c57%2c52%2c22%2c20%2c19%2c58%2c53%2c&StatusIDs=1%2c"
)

# Support searching by SchoolBranch (LOCATION_ID) — expects a 4-digit, zero-padded code
SCHOOLBRANCH_SEARCH_TEMPLATE = (
    "http://www.edna.pa.gov/Screens/wfSearchEntityResults.aspx?"
    "AUN=&SchoolBranch={BRANCH}&CurrentName=&City=&HistoricalName=&IU=-1&CID=-1&"
    "CategoryIDs=3%2c4%2c2%2c1%2c6%2c7%2c29%2c27%2c28%2c55%2c30%2c31%2c14%2c11%2c12%2c9%2c17%2c15%2c18%2c"
    "46%2c47%2c40%2c34%2c56%2c45%2c36%2c44%2c35%2c38%2c59%2c999%2c32%2c33%2c37%2c49%2c57%2c52%2c22%2c20%2c19%2c58%2c53%2c"
    "3%2c4%2c2%2c1%2c6%2c7%2c29%2c27%2c28%2c55%2c30%2c31%2c14%2c11%2c12%2c9%2c17%2c15%2c18%2c46%2c47%2c40%2c34%2c56%2c45%2c36%2c44%2c35%2c38%2c59%2c999%2c32%2c33%2c37%2c49%2c57%2c52%2c22%2c20%2c19%2c58%2c53%2c&StatusIDs=1%2c"
)

def _schoolbranch_search_url(branch_code: str) -> str:
    branch_code = (branch_code or "").strip()
    branch_code = re.sub(r"\D+", "", branch_code)
    branch_code = branch_code.zfill(4) if branch_code else ""
    return SCHOOLBRANCH_SEARCH_TEMPLATE.format(BRANCH=quote_plus(branch_code))

def _search_schoolbranch(session: requests.Session, branch_code: str) -> Tuple[List[Tuple[str, str, str]], str]:
    """
    Perform an EDNA search with SchoolBranch=<branch_code> (expects 4-digit string).
    Returns (candidates, search_url) where candidates is a list of (inst_name, branch, href).
    """
    url = _schoolbranch_search_url(branch_code)
    r = session.get(url); r.raise_for_status()
    soup = BeautifulSoup(r.text, "html.parser")
    table, inst_idx, branch_idx = _find_results_table_and_institution_col(soup)
    if not table or inst_idx is None:
        return [], url
    return _iter_institution_links(table, inst_idx, branch_idx), url

# Matches ONLY "javascript:__doPostBack(...)" exactly (keep if you still need it)
POSTBACK_RE = re.compile(
    r"""^javascript:\s*__doPostBack\(\s*'([^']*)'\s*,\s*'([^']*)'\s*\)\s*;?\s*$"""
)

# NEW: matches ANY occurrence of __doPostBack('...','...') inside href/onclick/etc.
POSTBACK_ANY_RE = re.compile(
    r"""__doPostBack\(\s*'(?P<target>[^']+)'\s*,\s*'(?P<arg>[^']*)'\s*\)"""
)

def _make_session() -> requests.Session:
    s = requests.Session()
    s.headers.update({
        "User-Agent": "Mozilla/5.0 (compatible; edna-lookup/1.0)",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.7",
        "Connection": "close",
    })
    orig = s.request

    def wrapped(method, url, **kwargs):
        kwargs.setdefault("timeout", (10, 30))
        return orig(method, url, **kwargs)

    s.request = wrapped
    return s

def _digits_only(s: str) -> str:
    return re.sub(r"\D+", "", s or "")

def _normalize_key(s: str) -> str:
    # Lowercase + collapse spaces + strip; longer limit for safety
    if not s:
        return ""
    s_norm = re.sub(r"\s+", " ", str(s)).strip().lower()
    return s_norm[:120]

def _pair_key(school: str, district: str) -> str:
    return f"{_normalize_key(school)}||{_normalize_key(district)}"

def _derive_district7_from_12(nces12: str) -> str:
    d = _digits_only(nces12)
    return d[:7] if len(d) >= 7 else ""

def _append_if_new(web_row: dict):
    """
    Maintain edna_output.csv with key = (School Name, District Name).

    Rules:
      - If no matching key exists -> append web_row.
      - If matching rows exist AND any have Status=='Open':
          * Do NOT replace the row(s).
          * Instead, non-destructively PATCH missing fields in-place from web_row
            (e.g., fill blank 'Grades', 'NCES Code', 'Detail URL', etc.).
      - Else (all matches Closed or blank Status) -> replace those rows with web_row.

    Columns are ensured and created if absent. Incoming Status is canonicalized.
    """
    append_cols = [
        "School Name",
        "School/Branch",
        "NCES Code",
        "Grades",
        "Detail URL",
        "District Name",
        "District NCES",
        "NCES 12-digit (District+Branch)",
        "Status",
    ]
    # Ensure incoming row has all keys
    for c in append_cols:
        web_row.setdefault(c, "")

    # Canonicalize Status
    web_row["Status"] = _status_canonicalize(web_row.get("Status", ""))

    # Ensure file exists with the right columns
    _ensure_csv_with_headers(EDNA_CACHE_CSV, append_cols)

    # Load
    existing = pd.read_csv(EDNA_CACHE_CSV, dtype=str).fillna("")
    for c in append_cols:
        if c not in existing.columns:
            existing[c] = ""

    # Match key
    in_school   = _normalize_key(web_row["School Name"])
    in_district = _normalize_key(web_row["District Name"])
    mask = (existing["School Name"].map(_normalize_key) == in_school) & \
           (existing["District Name"].map(_normalize_key) == in_district)
    matches = existing[mask]

    if matches.empty:
        _append_row_to_csv(EDNA_CACHE_CSV, append_cols, web_row)
        print(f"[ONLINE] Appended new row: {web_row['School Name']} / {web_row['District Name']} to {EDNA_CACHE_CSV}")
        return

    # We have at least one match
    any_open = any(_status_norm(s) == "open" for s in matches["Status"].tolist())

    if any_open:
        # Non-destructive PATCH: only fill blanks. Never overwrite nonblank values.
        patched = existing.copy()

        # Columns we are willing to patch if blank
        patchable = [
            "Grades",                       # ← this is the key addition you need
            "NCES Code",
            "School/Branch",
            "Detail URL",
            "District NCES",
            "NCES 12-digit (District+Branch)",
            "Status",                       # allow upgrading blank -> Open/Closed canonical
        ]

        for idx in patched[mask].index.tolist():
            for col in patchable:
                cur = str(patched.at[idx, col] or "").strip()
                newv = str(web_row.get(col, "") or "").strip()
                # fill only if current value is blank AND incoming has a value
                if (not cur) and newv:
                    patched.at[idx, col] = newv

        patched.to_csv(EDNA_CACHE_CSV, index=False)
        print(f"[ONLINE] Patched existing Open row(s) for {web_row['School Name']} / {web_row['District Name']} (filled missing fields where available)")
        return

    # All existing Status are Closed or blank → replace them with web_row (single canonical row)
    remaining = existing[~mask].copy()
    updated = pd.concat([remaining, pd.DataFrame([web_row])[append_cols]], ignore_index=True)
    updated.to_csv(EDNA_CACHE_CSV, index=False)
    print(f"[ONLINE] Replaced Closed/blank row(s) for {web_row['School Name']} / {web_row['District Name']} in {EDNA_CACHE_CSV}")

def _normalize_space(s: str) -> str:
    return re.sub(r"\s+", " ", s or "").strip()

def _currentname_search_url(school_name: str) -> str:
    return CURRENTNAME_SEARCH_TEMPLATE.format(CURRENT=quote_plus((school_name or "").strip()))

def _parse_postback_href(href: str) -> Optional[Tuple[str, str]]:
    """
    Parse an ASP.NET __doPostBack(...) from either a strict javascript: href
    or from an onclick (or any attribute) that merely contains the call.
    Returns (target, argument) or None.
    """
    if not href:
        return None
    val = href.strip()
    # Strict: javascript:__doPostBack('...','...')
    m = POSTBACK_RE.match(val)
    if m:
        return html.unescape(m.group(1)), html.unescape(m.group(2))
    # Flexible: find __doPostBack(...) anywhere (works for onclick handlers too)
    m2 = POSTBACK_ANY_RE.search(val)
    if m2:
        return html.unescape(m2.group("target")), html.unescape(m2.group("arg"))
    return None

def _collect_form_fields(form_tag: Tag) -> dict:
    data = {}
    for inp in form_tag.find_all("input"):
        name = inp.get("name")
        if name: data[name] = inp.get("value", "")
    for ta in form_tag.find_all("textarea"):
        name = ta.get("name")
        if name: data[name] = ta.get_text()
    for sel in form_tag.find_all("select"):
        name = sel.get("name")
        if not name: continue
        val = None
        for opt in sel.find_all("option"):
            if "selected" in opt.attrs:
                val = opt.get("value", opt.get_text()); break
        if val is None:
            first = sel.find("option")
            val = first.get("value", first.get_text()) if first else ""
        data[name] = val
    return data

def _do_postback(session: requests.Session, page_url: str, target: str, argument: str) -> Optional[BeautifulSoup]:
    """
    Perform an ASP.NET __doPostBack and return a BeautifulSoup of the *rendered content*.
    Handles both full-page HTML and Microsoft AJAX UpdatePanel delta responses.
    """
    # 1) GET the page so we can collect VIEWSTATE / EVENTVALIDATION fields
    r = session.get(page_url)
    r.raise_for_status()
    soup = BeautifulSoup(r.text, "html.parser")
    form = soup.find("form")
    if not form:
        print("[postback] no <form> on page to post back to")
        return None

    # 2) Build the POST payload with hidden fields + target/argument
    action = urljoin(page_url, form.get("action") or page_url)
    data = _collect_form_fields(form)
    data["__EVENTTARGET"] = target
    data["__EVENTARGUMENT"] = argument

    # 3) Post back. ASP.NET UpdatePanel often returns text/plain “delta” format.
    pr = session.post(
        action,
        data=data,
        headers={
            "Referer": page_url,
            "X-MicrosoftAjax": "Delta=true",
            "X-Requested-With": "XMLHttpRequest",
        },
    )
    pr.raise_for_status()

    text = pr.text or ""
    ctype = pr.headers.get("Content-Type", "").lower()
    if ("text/plain" in ctype or "application/json" in ctype or "|updatepanel|" in text.lower() or "|pageRedirect|" in text) and "|" in text:
        parts = text.split("|")
        html_frags = [p for p in parts if "<" in p and ">" in p]
        pick = None
        for frag in html_frags:
            if "<table" in frag.lower():
                pick = frag
                break
        if not pick and html_frags:
            pick = max(html_frags, key=len)
        if pick:
            bs = BeautifulSoup(pick, "html.parser")
            print(f"[postback] delta payload detected: picked fragment len={len(pick)} tables={len(bs.find_all('table'))}", flush=True)
            return bs
        print("[postback] delta payload had no obvious HTML fragment; returning raw parse", flush=True)
        return BeautifulSoup(text, "html.parser")

    return BeautifulSoup(text, "html.parser")

def _find_results_table_and_institution_col(soup: BeautifulSoup) -> Tuple[Optional[Tag], Optional[int], Optional[int]]:
    for table in soup.find_all("table"):
        for tr in table.find_all("tr"):
            ths = tr.find_all(["th", "td"])
            if not ths: continue
            hdrs = [_normalize_space(th.get_text(" ", strip=True)) for th in ths]
            inst_idx = branch_idx = None
            for i, h in enumerate(hdrs):
                hl = h.lower()
                if hl == "institution name": inst_idx = i
                if hl == "school/branch": branch_idx = i
            if inst_idx is not None:
                return table, inst_idx, branch_idx
    return None, None, None

def _iter_institution_links(table: Tag, inst_col_idx: int, branch_col_idx: Optional[int]) -> List[Tuple[str, str, str]]:
    out = []
    tbody = table.find("tbody") or table
    for tr in tbody.find_all("tr"):
        if tr.find("th"): continue
        tds = tr.find_all("td")
        if not tds or inst_col_idx >= len(tds): continue
        inst_cell = tds[inst_col_idx]
        a = inst_cell.find("a", href=True)
        if not a: continue
        href = a["href"].strip()
        name = _normalize_space(a.get_text(" ", strip=True))
        branch = ""
        if branch_col_idx is not None and branch_col_idx < len(tds):
            branch = _normalize_space(tds[branch_col_idx].get_text(" ", strip=True))
        out.append((name, branch, href))
    return out

def _find_table_with_header(soup: BeautifulSoup, header_name: str) -> Tuple[Optional[Tag], Optional[int], Optional[Tag]]:
    target = _normalize_space(header_name).lower()
    for table in soup.find_all("table"):
        for tr in table.find_all("tr", recursive=True):
            headers = tr.find_all("th") or tr.find_all("td")
            if not headers:
                continue
            hdrs = [_normalize_space(h.get_text(" ", strip=True)) for h in headers]
            for i, h in enumerate(hdrs):
                if _normalize_space(h).lower() == target:
                    return table, i, tr
    return None, None, None

def _extract_cell_below_header(soup: BeautifulSoup, header: str) -> str:
    table, col_idx, header_row = _find_table_with_header(soup, header)
    if not table or col_idx is None:
        return ""
    tbody = table.find("tbody")
    rows = []
    if tbody:
        rows = [tr for tr in tbody.find_all("tr") if tr.find_all("td")]
    else:
        for tr in header_row.find_all_next("tr"):
            if tr.find_parent("table") != table:
                break
            if tr.find_all("td"):
                rows.append(tr)
    if not rows:
        return ""
    tds = rows[0].find_all("td")
    if col_idx < len(tds):
        return _normalize_space(tds[col_idx].get_text(" ", strip=True))
    return ""

# ---------- Generic key–value extraction from detail page ----------
def _extract_kv_from_all_tables(soup: BeautifulSoup) -> dict:
    """
    Build a label->value dict by scanning every table and interpreting each row
    as alternating label/value pairs: [Label, Value, Label, Value, ...].
    Keys are returned exactly as shown on the page.
    """
    kv = {}
    for table in soup.find_all("table"):
        for tr in table.find_all("tr"):
            cells = tr.find_all(["th", "td"])
            if len(cells) < 2:
                continue
            texts = [_normalize_space(c.get_text(" ", strip=True)) for c in cells]
            for i in range(0, len(texts) - 1, 2):
                label = texts[i]
                value = texts[i + 1]
                if label:
                    kv[label] = value
    return kv

# ---------- Rewritten getters using the kv-dict ----------
def _school7_to_school5(school7: str) -> str:
    """
    Convert a 7-digit school NCES to the 5-digit 'SCH' component:
      * strip all non-digits
      * strip leading zeros until length ≤ 5
      * then left-pad with zeros to length 5
    If (after stripping non-digits) the string is empty, return ''.
    If the nonzero part still exceeds 5 (unexpected), take the last 5 and log a breadcrumb.
    """
    digits = re.sub(r"\D+", "", school7 or "")
    if not digits:
        return ""
    trimmed = digits.lstrip("0")
    if trimmed == "":
        return "00000"
    if len(trimmed) <= 5:
        return trimmed.zfill(5)
    print(f"[nces][warn] school7 '{school7}' reduced to >5='{trimmed}', using last 5")
    return trimmed[-5:]

def _extract_school_nces7_from_details(soup: BeautifulSoup) -> str:
    """
    Return a 7-digit School NCES code if present; else ''.
    Strategy:
      1) KV-style labels that commonly carry the school NCES (Entity tab)
      2) Demographics header cell under 'NCES Code'
      3) Proximity regex: a 7-digit number within ±100 chars of 'NCES'
    Only a *7-digit* sequence is returned from here.
    """
    def _first_7(s: str) -> str:
        if not s:
            return ""
        m = re.search(r"\b\d{7}\b", s)
        return m.group(0) if m else ""

    kv = _extract_kv_from_all_tables(soup)
    for label in ("NCES Code", "School NCES", "NCES School Code"):
        cand = kv.get(label)
        n7 = _first_7(cand or "")
        if n7:
            return n7

    demog = _extract_cell_below_header(soup, "NCES Code")
    n7 = _first_7(demog or "")
    if n7:
        return n7

    text = soup.get_text(" ", strip=True)
    for m in re.finditer(r"\b\d{7}\b", text):
        start, end = m.start(), m.end()
        window = text[max(0, start-100):min(len(text), end+100)].lower()
        if "nces" in window:
            return m.group(0)

    return ""

def _extract_district_name_from_details(soup: BeautifulSoup) -> str:
    """
    District/LEA name: usually 'LEA Name' on the Entity tab.
    """
    kv = _extract_kv_from_all_tables(soup)
    for label in ("LEA Name", "District Name", "School District", "LEA"):
        if kv.get(label):
            return _normalize_space(kv[label])
    return ""

def _extract_district_nces_from_details(soup: BeautifulSoup) -> str:
    """
    District NCES on a **district page**:
      1) KV labels if present (rare)
      2) Demographics header table under 'NCES Code' (common)
    Returns digits-only string or ''.
    """
    kv = _extract_kv_from_all_tables(soup)
    for label in ("District NCES", "NCES District Code", "District NCES Code", "LEA NCES"):
        if kv.get(label):
            return _digits_only(kv[label])

    return _digits_only(_extract_cell_below_header(soup, "NCES Code"))

def _normalize_strict(s: str) -> str:
    """
    Normalize a string for strict comparison:
    - Convert to string
    - Lowercase
    - Collapse internal whitespace
    - Strip leading/trailing whitespace
    - Truncate to the first 40 characters
    """
    if not s:
        return ""
    s_norm = re.sub(r"\s+", " ", str(s)).strip().lower()
    return s_norm[:40]

def _district_equals(a: str, b: str) -> bool:
    """Require strict equality after normalization (with 40-char truncation)."""
    return _normalize_strict(a) == _normalize_strict(b)

def _search_currentname(session: requests.Session, current_name: str) -> Tuple[List[Tuple[str, str, str]], str]:
    """
    Perform an EDNA search with CurrentName=<current_name>.
    Returns (candidates, search_url) where candidates is a list of (inst_name, branch, href).
    """
    url = _currentname_search_url(current_name)
    r = session.get(url); r.raise_for_status()
    soup = BeautifulSoup(r.text, "html.parser")
    table, inst_idx, branch_idx = _find_results_table_and_institution_col(soup)
    if not table or inst_idx is None:
        return [], url
    return _iter_institution_links(table, inst_idx, branch_idx), url

def _fetch_detail_soup(session: requests.Session, search_url: str, href: str) -> Tuple[Optional[BeautifulSoup], str]:
    """
    Follow either a direct link or an ASP.NET postback from a search results row
    and return (detail_soup, detail_url_for_csv).
    """
    try:
        if href.lower().startswith("javascript:"):
            parsed = _parse_postback_href(href)
            if not parsed:
                return None, ""
            target, argument = parsed
            soup = _do_postback(session, search_url, target, argument)
            if soup is not None:
                # For delta postbacks we won't have a stable canonical URL; retain search_url for context.
                soup.base_url = search_url
            return soup, ""
        else:
            norm_href = _normalize_detail_url(href)
            detail_url = urljoin(EDNA_BASE, norm_href)
            r = session.get(detail_url)
            r.raise_for_status()
            bs = BeautifulSoup(r.text, "html.parser")
            bs.base_url = detail_url
            return bs, detail_url
    except Exception as e:
        print(f"[fetch-detail] {e}")
        traceback.print_exc()
        return None, ""

# ---------- Grades extraction (ported & adapted from aun_to_nces.py) ----------
def _table_header_texts(tr: Tag) -> list:
    headers = tr.find_all("th") or tr.find_all("td")
    return [_normalize_space(h.get_text(" ", strip=True)) for h in headers]

def _find_table_with_header_loose(soup: BeautifulSoup, header_name: str) -> Tuple[Optional[Tag], Optional[int], Optional[Tag]]:
    """Loose finder used by grades extractor; similar to above helper."""
    for table in soup.find_all("table"):
        header_row = None
        for tr in table.find_all("tr", recursive=True):
            if tr.find("th") or tr.find_all("td"):
                header_row = tr
                break
        if not header_row:
            continue
        headers = _table_header_texts(header_row)
        for idx, h in enumerate(headers):
            if _normalize_space(h).lower() == _normalize_space(header_name).lower():
                return table, idx, header_row
    return None, None, None

def _extract_grades_from_details(soup: BeautifulSoup) -> str:
    """
    Return a canonical comma-separated grade band (e.g., 'PK, K, 1, 2, ...').
    Looks for:
      1) a table headed 'Grades' (original behavior), OR
      2) any key-value row whose label mentions 'grade' (e.g., 'Grades Served', 'Grade Span').
    Understands ranges like 'K-12', 'PK–K', '7–12'. Silent (no prints).
    """
    try:
        def _canon_token(t: str) -> str:
            t = t.strip().upper()
            if t in ("PK", "K"):
                return t
            m = re.fullmatch(r"\d{1,2}", t)
            return str(int(t)) if m else ""  # normalize leading zeros

        def _rank(t: str) -> int:
            if t == "PK": return 0
            if t == "K":  return 1
            if re.fullmatch(r"\d{1,2}", t): return int(t) + 1
            return 999

        def _expand_span(a: str, b: str) -> list[str]:
            A, B = _canon_token(a), _canon_token(b)
            if not A or not B:
                return []
            order = ["PK", "K"] + [str(i) for i in range(1, 13)]
            ia, ib = order.index(A) if A in order else -1, order.index(B) if B in order else -1
            if ia < 0 or ib < 0:
                return []
            if ia > ib:
                ia, ib = ib, ia
            return order[ia:ib+1]

        def _tokens_from_text(txt: str) -> list[str]:
            if not txt:
                return []
            s = _normalize_space(txt)
            # Split on commas/semicolons first
            parts = re.split(r"[;,]", s)
            tokens: list[str] = []
            for part in parts:
                # handle spans with -, –, — (hyphen/en/em dashes)
                m = re.search(r"\b(PK|K|\d{1,2})\s*[-–—]\s*(PK|K|\d{1,2})\b", part, flags=re.IGNORECASE)
                if m:
                    tokens += _expand_span(m.group(1), m.group(2))
                # pick up isolated tokens too
                for t in re.findall(r"\b(PK|K|\d{1,2})\b", part, flags=re.IGNORECASE):
                    ct = _canon_token(t)
                    if ct:
                        tokens.append(ct)
            # dedupe & sort semantically
            seen, uniq = set(), []
            for t in tokens:
                if t and t not in seen:
                    seen.add(t); uniq.append(t)
            uniq.sort(key=_rank)
            # collapse to canonical ordered set
            order = ["PK", "K"] + [str(i) for i in range(1, 13)]
            return [t for t in order if t in uniq]

        # --- Path 1: original table headed 'Grades'
        table, grades_col_idx, header_row = _find_table_with_header_loose(soup, "Grades")
        if table:
            tbody = table.find("tbody")
            data_rows = []
            if tbody:
                for tr in tbody.find_all("tr"):
                    if tr.find("th"): continue
                    if tr.find_all("td"): data_rows.append(tr)
            else:
                if header_row:
                    for tr in header_row.find_all_next("tr"):
                        if tr.find_parent("table") != table: break
                        if tr.find("th"): continue
                        if tr.find_all("td"): data_rows.append(tr)
                else:
                    for tr in table.find_all("tr", recursive=True):
                        if tr.find_all("td"): data_rows.append(tr)

            texts = []
            for tr in data_rows:
                tds = tr.find_all("td")
                if not tds: continue
                if grades_col_idx is not None and grades_col_idx < len(tds):
                    texts.append(_normalize_space(tds[grades_col_idx].get_text(" ", strip=True)))
                else:
                    texts.append(_normalize_space(tr.get_text(" ", strip=True)))

            tokens = []
            for txt in texts:
                tokens += _tokens_from_text(txt)

            if tokens:
                return ", ".join(tokens)

        # --- Path 2: KV labels that mention 'grade'
        kv = _extract_kv_from_all_tables(soup)
        # prefer more specific keys first
        preferred_labels = [
            "Grades Served", "Grades Offered", "Grade Span", "Grades",
            "Lowest Grade", "Highest Grade",
        ]
        # 2a) direct preferred labels
        for label in preferred_labels:
            if label in kv and kv[label]:
                toks = _tokens_from_text(kv[label])
                if toks:
                    return ", ".join(toks)

        # 2b) any label containing 'grade'
        for label, value in kv.items():
            if "grade" in label.lower() and value:
                toks = _tokens_from_text(value)
                if toks:
                    return ", ".join(toks)

        return ""
    except Exception:
        traceback.print_exc()
        return ""

# ---- Grade min/max helpers (derived from the 'Grades' band) ----
_GRADE_RANK = {"PK": 0, "K": 1}
_GRADE_RANK.update({str(i): i + 1 for i in range(1, 13)})  # 1..12 → 2..13

def _parse_grade_tokens(grades_str: str) -> list[str]:
    """
    Parse a 'Grades' string like 'PK, K, 1, 2, 3, 4' into canonical tokens.
    Recognizes PK, K, and numerals (1..12).
    Returns tokens in input order, deduped.
    """
    if not grades_str:
        return []
    tokens = re.findall(r"\b(?:PK|K|\d{1,2})\b", grades_str, flags=re.IGNORECASE)
    canon = []
    seen = set()
    for t in tokens:
        t_up = t.upper()
        # Canonicalize numbers (e.g., '01' → '1')
        if t_up not in ("PK", "K"):
            t_up = str(int(t_up))  # safe due to regex
        if t_up not in seen:
            seen.add(t_up)
            canon.append(t_up)
    return canon

def _lowest_highest_from_tokens(tokens: list[str]) -> tuple[str, str]:
    """
    Compute the lowest and highest grade label given canonical tokens.
    Returns ("", "") if tokens empty or none are recognized.
    """
    ranked = [(t, _GRADE_RANK[t]) for t in tokens if t in _GRADE_RANK]
    if not ranked:
        return "", ""
    ranked.sort(key=lambda x: x[1])
    return ranked[0][0], ranked[-1][0]

def edna_lookup_district_by_name(session: requests.Session, lea_name: str, delay_sec: float = 0.6) -> Tuple[str, str]:
    candidates, search_url = _search_currentname(session, lea_name)
    if not candidates:
        print(f"[online-lookup][district] no candidates for '{lea_name}'")
        return "", ""

    target_norm = _normalize_strict(lea_name)
    for i, (inst_name, _branch, href) in enumerate(candidates, start=1):
        if _normalize_strict(inst_name) != target_norm:
            continue
        try:
            time.sleep(delay_sec)
            detail_soup, detail_url = _fetch_detail_soup(session, search_url, href)
        except Exception as e:
            print(f"[online-lookup][district] candidate #{i} fetch failed: {e}")
            continue
        if not detail_soup:
            continue

        district_nces_digits = _extract_district_nces_from_details(detail_soup)
        if len(district_nces_digits) == 7:
            print(f"[online-lookup][district] accepted '{inst_name}' → NCES {district_nces_digits}")
            return district_nces_digits, detail_url

        print(f"[online-lookup][district] '{inst_name}' has no 7-digit NCES on page (saw: '{district_nces_digits}')")
    print(f"[online-lookup][district] no exact Institution Name match for '{lea_name}'")
    return "", ""

def edna_lookup_by_name(school_name: str, expected_district: str, delay_sec: float = 0.6) -> Optional[dict]:
    school_name = norm(school_name)
    expected_district = norm(expected_district)
    if not school_name:
        return None

    session = _make_session()
    try:
        candidates, search_url = _search_currentname(session, school_name)
    except Exception:
        traceback.print_exc()
        print(f"[grades] url={_force_screens_url(_currentname_search_url(school_name))} school=\"{school_name}\" district=\"{expected_district}\" branch=— grades=\"\"")
        return None

    if not candidates:
        print(f"[grades] url={_force_screens_url(_currentname_search_url(school_name))} school=\"{school_name}\" district=\"{expected_district}\" branch=— grades=\"—\"")
        return None

    for i, (inst_name, branch, href) in enumerate(candidates, start=1):
        if _normalize_strict(inst_name) != _normalize_strict(school_name):
            continue
        try:
            time.sleep(delay_sec)
            # --- Force /Screens/ here ---
            detail_soup, detail_url_for_csv = _fetch_detail_soup(session, search_url, _force_screens_url(href))
        except Exception:
            traceback.print_exc()
            print(f"[grades] url={_force_screens_url(search_url)} school=\"{inst_name}\" district=\"{expected_district}\" branch={branch or '—'} grades=\"\"")
            continue
        if not detail_soup:
            print(f"[grades] url={_force_screens_url(search_url)} school=\"{inst_name}\" district=\"{expected_district}\" branch={branch or '—'} grades=\"—\"")
            continue

        school7 = _extract_school_nces7_from_details(detail_soup) or ""
        district_name = _extract_district_name_from_details(detail_soup) or ""
        grades = _extract_grades_from_details(detail_soup) or ""
        url_for_log = _force_screens_url(detail_url_for_csv or getattr(detail_soup, "base_url", "") or search_url)

        if not _district_equals(district_name, expected_district):
            print(f"[grades] url={url_for_log} school=\"{inst_name}\" district=\"{district_name or '—'}\" branch={branch or '—'} grades=\"{grades or '—'}\"")
            continue

        district7, _ = edna_lookup_district_by_name(session, district_name, delay_sec)
        school5 = _school7_to_school5(school7) if school7 else ""
        nces12 = f"{district7}{school5}" if (len(district7) == 7 and len(school5) == 5) else ""

        row = {
            "School Name": inst_name,
            "School/Branch": f'="{branch}"' if branch else "",
            "NCES Code": f'="{school7}"' if school7 else "",
            "Grades": grades,
            "Detail URL": url_for_log,  # already normalized
            "District Name": district_name,
            "District NCES": f'="{district7}"' if district7 else "",
            "NCES 12-digit (District+Branch)": f'="{nces12}"' if nces12 else "",
        }
        print(f"[grades] url={url_for_log} school=\"{inst_name}\" district=\"{district_name}\" branch={branch or '—'} grades=\"{grades or '—'}\"")
        return row

    print(f"[grades] url={_force_screens_url(_currentname_search_url(school_name))} school=\"{school_name}\" district=\"{expected_district}\" branch=— grades=\"—\"")
    return None

def edna_lookup_by_location_id(location_id: str, delay_sec: float = 0.6) -> Optional[dict]:
    """
    Online lookup by SchoolBranch=<LOCATION_ID>.
    Prints exactly one [grades] line for the accepted candidate; otherwise logs with empty grades.
    """
    loc = _digits_only(location_id).zfill(4)
    if not loc or len(loc) != 4:
        print(f"[grades] url={EDNA_BASE}/Screens/wfSearchEntityResults.aspx school=\"—\" district=\"—\" branch={loc or '—'} grades=\"—\"")
        return None

    session = _make_session()
    try:
        candidates, search_url = _search_schoolbranch(session, loc)
    except Exception:
        traceback.print_exc()
        print(f"[grades] url={_schoolbranch_search_url(loc)} school=\"—\" district=\"—\" branch={loc} grades=\"\"")
        return None

    if not candidates:
        print(f"[grades] url={_schoolbranch_search_url(loc)} school=\"—\" district=\"—\" branch={loc} grades=\"—\"")
        return None

    for i, (inst_name, branch, href) in enumerate(candidates, start=1):
        try:
            time.sleep(delay_sec)
            detail_soup, detail_url_for_csv = _fetch_detail_soup(session, search_url, href)
        except Exception:
            traceback.print_exc()
            print(f"[grades] url={search_url} school=\"{inst_name}\" district=\"—\" branch={branch or '—'} grades=\"\"")
            continue
        if not detail_soup:
            print(f"[grades] url={search_url} school=\"{inst_name}\" district=\"—\" branch={branch or '—'} grades=\"—\"")
            continue

        school7 = _extract_school_nces7_from_details(detail_soup) or ""
        district_name = _extract_district_name_from_details(detail_soup) or ""
        grades = _extract_grades_from_details(detail_soup) or ""
        url_for_log = detail_url_for_csv or getattr(detail_soup, "base_url", "") or search_url

        district7, _ = edna_lookup_district_by_name(session, district_name, delay_sec)
        school5 = _school7_to_school5(school7) if school7 else ""
        nces12 = f"{district7}{school5}" if (len(district7) == 7 and len(school5) == 5) else ""

        if not nces12:
            print(f"[grades] url={url_for_log} school=\"{inst_name}\" district=\"{district_name or '—'}\" branch={branch or '—'} grades=\"{grades or '—'}\"")
            continue

        row = {
            "School Name": inst_name,
            "School/Branch": f'="{branch}"' if branch else "",
            "NCES Code": f'="{school7}"' if school7 else "",
            "Grades": grades,
            "Detail URL": detail_url_for_csv,
            "District Name": district_name,
            "District NCES": f'="{district7}"' if district7 else "",
            "NCES 12-digit (District+Branch)": f'="{nces12}"',
        }
        print(f"[grades] url={url_for_log} school=\"{inst_name}\" district=\"{district_name or '—'}\" branch={branch or '—'} grades=\"{grades or '—'}\"")
        return row

    print(f"[grades] url={_schoolbranch_search_url(loc)} school=\"—\" district=\"—\" branch={loc} grades=\"—\"")
    return None

# ==============================
# Edna District pre-crawl: collect schools into edna_output.csv
# ==============================

def _ensure_csv_with_headers(path: str, cols: list[str]):
    """Create CSV with given columns if it doesn't exist."""
    if not os.path.exists(path):
        pd.DataFrame(columns=cols).to_csv(path, index=False)

def _append_row_to_csv(path: str, cols: list[str], row: dict):
    """Append a single row (dict) to CSV, creating file with headers if needed."""
    _ensure_csv_with_headers(path, cols)
    pd.DataFrame([row])[cols].to_csv(path, mode="a", index=False, header=False)

def _ensure_output_csv_exists():
    """Guarantee that edna_output.csv exists with the proper columns (including Grades)."""
    cols = [
        "School Name",
        "School/Branch",
        "NCES Code",
        "Grades",  # NEW
        "Detail URL",
        "District Name",
        "District NCES",
        "NCES 12-digit (District+Branch)",
        "Status",
    ]
    _ensure_csv_with_headers(EDNA_CACHE_CSV, cols)

def _collect_unique_district_names_from_workbook(xlsx_path: str, sheets: list[str]) -> list[str]:
    """
    Load the specified sheets and return a de-duplicated, normalized list of District Names.
    """
    districts = []
    for sheet in sheets:
        try:
            df = pd.read_excel(xlsx_path, sheet_name=sheet, dtype=str).fillna("")
            df.rename(columns=lambda c: c.strip() if isinstance(c, str) else c, inplace=True)
            if "District Name" in df.columns:
                districts.extend([norm(x) for x in df["District Name"].tolist() if norm(x)])
            else:
                print(f"[district-prep] Sheet '{sheet}' missing 'District Name' column")
        except Exception as e:
            print(f"[district-prep] Failed reading sheet '{sheet}': {e}")
            traceback.print_exc()
    seen = set()
    out = []
    for d in districts:
        k = _normalize_strict(d)
        if k and k not in seen:
            seen.add(k)
            out.append(d)
    return out

def _click_link_or_postback(session: requests.Session, page_url: str, a_tag: Tag) -> Optional[BeautifulSoup]:
    href = (a_tag.get("href") or "").strip()
    if not href:
        print("[nav] anchor missing href")
        return None

    def _fetch(url: str, ref: str = "") -> Optional[BeautifulSoup]:
        try:
            r = session.get(url, headers={"Referer": ref} if ref else None)
            r.raise_for_status()
            bs = BeautifulSoup(r.text, "html.parser")
            bs.base_url = url
            return bs
        except Exception as e:
            print(f"[nav] GET {url} failed: {e}")
            traceback.print_exc()
            return None

    # Handle postbacks
    if href.lower().startswith("javascript:"):
        parsed = _parse_postback_href(href)
        if not parsed:
            print(f"[postback] could not parse postback href: {href[:120]}...")
            return None
        target, argument = parsed
        try:
            soup = _do_postback(session, page_url, target, argument)
        except Exception as e:
            print(f"[postback] error: {e}")
            traceback.print_exc()
            return None
        nav_url = page_url
    else:
        # <<< normalize detail URLs here >>>
        norm_href = _normalize_detail_url(href)
        nav_url = urljoin(EDNA_BASE, norm_href)
        soup = _fetch(nav_url, ref=page_url)
        if not soup:
            return None

    if soup:
        print(f"[trace] after click/postback: url≈{nav_url} len(html)={len(str(soup))} tables={len(soup.find_all('table'))}", flush=True)

    # Follow likely school/branches iframe if present (unchanged)
    try:
        iframes = soup.find_all("iframe")
        if len(iframes) == 1 and iframes[0].get("src"):
            iframe_src = urljoin(nav_url, iframes[0].get("src"))
            print(f"[trace] following iframe src: {iframe_src}", flush=True)
            soup_iframe = _fetch(iframe_src, ref=nav_url)
            if soup_iframe:
                print(f"[trace] iframe fetch len(html)={len(str(soup_iframe))} tables={len(soup_iframe.find_all('table'))}", flush=True)
                return soup_iframe
        elif len(iframes) > 1:
            for fr in iframes:
                src = fr.get("src", "")
                title = fr.get("title", "")
                txt = (src + " " + title).lower()
                if "school" in txt or "branch" in txt:
                    iframe_src = urljoin(nav_url, src)
                    print(f"[trace] following probable schools iframe src: {iframe_src}", flush=True)
                    soup_iframe = _fetch(iframe_src, ref=nav_url)
                    if soup_iframe:
                        print(f"[trace] iframe fetch len(html)={len(str(soup_iframe))} tables={len(soup_iframe.find_all('table'))}", flush=True)
                        return soup_iframe
    except Exception as e:
        print(f"[trace] iframe follow error: {e}")
        traceback.print_exc()

    return soup

def _schools_table_parse(soup: BeautifulSoup) -> list[dict]:
    """
    Robust parser for a district's 'Schools/Branches' listing.
    Returns a list of dicts: {name, branch, nces7, status, detail_href}.
    NEVER returns None; returns [] if nothing parseable is found.
    """

    def canon(s: str) -> str:
        s = _normalize_space(s).lower()
        s = re.sub(r"[^a-z0-9]+", " ", s)
        return s.strip()

    def header_index_map(header_cells: list[str]) -> tuple[Optional[int], Optional[int], Optional[int], Optional[int]]:
        inst_idx = branch_idx = nces_idx = status_idx = None
        H = [canon(h) for h in header_cells]

        for i, h in enumerate(H):
            if "institution" in h and "name" in h: inst_idx = i; break
        if inst_idx is None:
            for i, h in enumerate(H):
                if "school" in h and "name" in h: inst_idx = i; break
        if inst_idx is None:
            for i, h in enumerate(H):
                if h == "name": inst_idx = i; break

        for i, h in enumerate(H):
            if "school" in h and "branch" in h: branch_idx = i; break
        if branch_idx is None:
            for i, h in enumerate(H):
                if h in {"branch", "branch code", "location id"} or ("branch" in h and "code" in h):
                    branch_idx = i; break

        for i, h in enumerate(H):
            if "nces" in h: nces_idx = i; break

        for i, h in enumerate(H):
            if "status" in h: status_idx = i; break

        return inst_idx, branch_idx, nces_idx, status_idx

    def extract_headers_from_table(table: Tag) -> list[str]:
        thead = table.find("thead")
        if thead:
            tr = thead.find("tr")
            if tr:
                return [_normalize_space(th.get_text(" ", strip=True)) for th in tr.find_all(["th", "td"])]
        tr = table.find("tr")
        if tr:
            return [_normalize_space(th.get_text(" ", strip=True)) for th in tr.find_all(["th", "td"])]
        return []

    results: list[dict] = []
    tables = soup.find_all("table")
    print(f"[trace] _schools_table_parse: found {len(tables)} table(s) on page", flush=True)

    # -------- Pass 1: header-driven extraction
    for table in tables:
        headers = extract_headers_from_table(table)
        if not headers:
            continue
        inst_idx, branch_idx, nces_idx, status_idx = header_index_map(headers)
        if inst_idx is None or branch_idx is None:
            continue

        body = table.find("tbody") or table
        row_count = 0
        for tr in body.find_all("tr"):
            if tr.find("th"):
                continue
            tds = tr.find_all("td")
            if not tds:
                continue
            mx = max(inst_idx, branch_idx)
            if len(tds) <= mx:
                continue

            name_cell = tds[inst_idx]
            name = _normalize_space(name_cell.get_text(" ", strip=True))
            a = name_cell.find("a", href=True)
            href_raw = a["href"].strip() if a else ""
            href = _force_screens_url(href_raw)  # <- normalize immediately

            branch = _normalize_space(tds[branch_idx].get_text(" ", strip=True))

            nces7 = ""
            if nces_idx is not None and nces_idx < len(tds):
                nces7 = _digits_only(tds[nces_idx].get_text(" ", strip=True))

            status = ""
            if status_idx is not None and status_idx < len(tds):
                status = _status_canonicalize(tds[status_idx].get_text(" ", strip=True))

            if name and branch:
                results.append({
                    "name": name,
                    "branch": branch,
                    "nces7": nces7,
                    "status": status,
                    "detail_href": href,
                })
                row_count += 1

        if row_count:
            print(f"[trace] header-parse hit: headers={headers} rows={row_count}", flush=True)
            return results  # List (non-empty)

    # -------- Pass 2: shape-driven fallback
    def looks_like_branch(s: str) -> bool:
        return bool(re.fullmatch(r"\d{4}", _digits_only(s).zfill(4)))

    for table in tables:
        body = table.find("tbody") or table
        for tr in body.find_all("tr"):
            tds = tr.find_all("td")
            if not tds:
                continue
            link_cell = None
            branch_cell_text = ""
            for td in tds:
                if (not link_cell) and td.find("a", href=True):
                    link_cell = td
                text = _normalize_space(td.get_text(" ", strip=True))
                if not branch_cell_text and looks_like_branch(text):
                    branch_cell_text = _digits_only(text).zfill(4)
            if link_cell and branch_cell_text:
                name = _normalize_space(link_cell.get_text(" ", strip=True))
                href_raw = link_cell.find("a", href=True)["href"].strip()
                href = _force_screens_url(href_raw)
                results.append({
                    "name": name,
                    "branch": branch_cell_text,
                    "nces7": "",
                    "status": "",
                    "detail_href": href,
                })

    if results:
        print(f"[trace] shape-parse hit: rows={len(results)} (no reliable headers)", flush=True)
        return results

    # -------- Nothing parseable; return an empty list (not None)
    hdr_summaries = []
    for t in tables[:3]:
        hdr_summaries.append(extract_headers_from_table(t))
    print(f"[trace] no rows parsed; sample headers={hdr_summaries}", flush=True)
    return []  

def _detect_current_page_num(pager_root: Tag) -> int:
    try:
        for sp in pager_root.find_all("span"):
            txt = _normalize_space(sp.get_text(" ", strip=True))
            if re.fullmatch(r"\d+", txt):
                sib_text = " ".join(a.get_text(" ", strip=True) for a in pager_root.find_all("a"))
                if re.search(r"\b\d+\b", sib_text):
                    return int(txt)
    except Exception:
        traceback.print_exc()
    return 0

def _paginate_next(session: requests.Session,
                   current_url: str,
                   current_soup: BeautifulSoup) -> Optional[tuple[BeautifulSoup, str]]:
    try:
        for a in current_soup.find_all("a", href=True):
            label = _normalize_space(a.get_text(" ", strip=True)).lower()
            if label in {"next", "next >", ">", "›", "»"}:
                soup2 = _click_link_or_postback(session, current_url, a)
                return (soup2, current_url) if soup2 else None

        target_next = None
        arg_next = None
        numeric_candidates: list[tuple[int, str, str, Tag]] = []

        def consider_element(tag: Tag):
            nonlocal target_next, arg_next, numeric_candidates
            for attr in ("href", "onclick"):
                val = (tag.get(attr) or "").strip()
                if not val:
                    continue
                m = POSTBACK_RE.match(val)
                if not m:
                    m2 = POSTBACK_ANY_RE.search(val)
                    if not m2:
                        continue
                    tgt, arg = html.unescape(m2.group("target")), html.unescape(m2.group("arg"))
                else:
                    tgt, arg = html.unescape(m.group(1)), html.unescape(m.group(2))

                if arg == "Page$Next":
                    target_next, arg_next = tgt, arg
                    return

                mn = re.match(r"^Page\$(\d+)$", arg)
                if mn:
                    try:
                        numeric_candidates.append((int(mn.group(1)), tgt, arg, tag))
                    except Exception:
                        pass

        for tag in current_soup.find_all(["a", "button", "span", "input"]):
            consider_element(tag)
            if target_next and arg_next:
                break

        if target_next and arg_next:
            soup2 = _do_postback(session, current_url, target_next, arg_next)
            return (soup2, current_url) if soup2 else None

        if numeric_candidates:
            pager_root = None
            try:
                parents = []
                for _, _, _, tag in numeric_candidates[:6]:
                    chain = []
                    p = tag
                    while p:
                        chain.append(p)
                        p = p.parent
                    parents.append(chain)
                common = []
                for cols in zip(*map(reversed, parents)):
                    if all(x is cols[0] for x in cols):
                        common.append(cols[0])
                    else:
                        break
                pager_root = common[-1] if common else None
            except Exception:
                traceback.print_exc()

            pager_root = pager_root or current_soup
            cur = _detect_current_page_num(pager_root)
            if cur > 0:
                for page_num, tgt, arg, _ in sorted(numeric_candidates, key=lambda t: t[0]):
                    if page_num == cur + 1:
                        soup2 = _do_postback(session, current_url, tgt, arg)
                        return (soup2, current_url) if soup2 else None

            for page_num, tgt, arg, _ in sorted(numeric_candidates, key=lambda t: t[0]):
                if page_num >= 2:
                    soup2 = _do_postback(session, current_url, tgt, arg)
                    return (soup2, current_url) if soup2 else None

        for tag in current_soup.find_all(True):
            labelled = (_normalize_space(tag.get("aria-label", "")) + " " +
                        _normalize_space(tag.get("title", ""))).lower()
            if any(k in labelled for k in ("next", "right", "forward")):
                for attr in ("href", "onclick"):
                    val = (tag.get(attr) or "").strip()
                    if not val:
                        continue
                    m = POSTBACK_ANY_RE.search(val) or POSTBACK_RE.match(val)
                    if m:
                        if hasattr(m, "groupdict") and "target" in m.groupdict():
                            tgt, arg = html.unescape(m.group("target")), html.unescape(m.group("arg"))
                        else:
                            tgt, arg = html.unescape(m.group(1)), html.unescape(m.group(2))
                        soup2 = _do_postback(session, current_url, tgt, arg)
                        return (soup2, current_url) if soup2 else None

        return None

    except Exception as e:
        print(f"[_paginate_next] {e}")
        traceback.print_exc()
        return None

def _open_schools_branches_direct(session: requests.Session,
                                  detail_soup: BeautifulSoup,
                                  referer_url: str = "") -> Optional[tuple[BeautifulSoup, str]]:
    for a in detail_soup.find_all("a", href=True):
        href = a["href"]
        if "wfSchools.aspx" in href and "ID=" in href:
            schools_url = urljoin(EDNA_BASE, href)
            try:
                r = session.get(schools_url, headers={"Referer": referer_url} if referer_url else None)
                r.raise_for_status()
                soup2 = BeautifulSoup(r.text, "html.parser")
                return soup2, schools_url
            except Exception as e:
                print(f"[district] direct schools link failed: {e}")
                traceback.print_exc()
                return None

    entity_id = _extract_entity_id_from_soup(detail_soup, fallback_urls=[referer_url])
    if not entity_id:
        print("[district] could not find entity ID to build wfSchools.aspx link")
        return None

    schools_url = _build_schools_url(entity_id)
    if not schools_url:
        print("[district] failed to build wfSchools.aspx URL")
        return None

    try:
        r = session.get(schools_url, headers={"Referer": referer_url} if referer_url else None)
        r.raise_for_status()
        soup2 = BeautifulSoup(r.text, "html.parser")
        return soup2, schools_url
    except Exception as e:
        print(f"[district] GET schools page failed: {e}")
        traceback.print_exc()
        return None

def _append_school_row_from_listing(district_name: str, district7: str,
                                    listing_row: dict, session: requests.Session, referer_url: str):
    school_name = listing_row.get("name", "")
    branch = listing_row.get("branch", "")
    school7 = listing_row.get("nces7", "")
    href = listing_row.get("detail_href", "")
    status = _status_canonicalize(listing_row.get("status", ""))

    grades = ""
    detail_url = ""

    try:
        if href and not href.lower().startswith("javascript:"):
            # --- Force /Screens/ here ---
            detail_url = _force_screens_url(href)
            r = session.get(detail_url, headers={"Referer": referer_url} if referer_url else None)
            r.raise_for_status()
            dsoup = BeautifulSoup(r.text, "html.parser"); dsoup.base_url = detail_url

            school7_page = _extract_school_nces7_from_details(dsoup) or ""
            if school7_page and not school7:
                school7 = school7_page

            grades = _extract_grades_from_details(dsoup) or ""

            # --- Force /Screens/ in the one-liner too ---
            print(f"[grades] url={_force_screens_url(detail_url)} school=\"{school_name}\" district=\"{district_name}\" branch={branch or '—'} grades=\"{grades or '—'}\"")
        else:
            url_for_log = referer_url or (EDNA_BASE + "/")
            print(f"[grades] url={_force_screens_url(url_for_log)} school=\"{school_name}\" district=\"{district_name}\" branch={branch or '—'} grades=\"—\"")
    except Exception:
        url_for_log = detail_url or referer_url or (EDNA_BASE + "/")
        print(f"[grades] url={_force_screens_url(url_for_log)} school=\"{school_name}\" district=\"{district_name}\" branch={branch or '—'} grades=\"\"")
        traceback.print_exc()

    school5 = _school7_to_school5(school7) if school7 else ""
    nces12 = f"{district7}{school5}" if (len(_digits_only(district7)) == 7 and len(school5) == 5) else ""

    web_row = {
        "School Name": school_name,
        "School/Branch": f'="{branch}"' if branch else "",
        "NCES Code": f'="{school7}"' if school7 else "",
        "Grades": grades,
        "Detail URL": detail_url,
        "District Name": district_name,
        "District NCES": f'="{district7}"' if district7 else "",
        "NCES 12-digit (District+Branch)": f'="{nces12}"' if nces12 else "",
        "Status": status,
    }

    try:
        _append_if_new(web_row)
    except Exception as e:
        print(f"[district-prepend] {_normalize_space(school_name)} append failed: {e}")
        traceback.print_exc()

def _extract_entity_id_from_soup(detail_soup: BeautifulSoup, fallback_urls: list[str] = None) -> str:
    for a in detail_soup.find_all("a", href=True):
        href = a["href"]
        m = re.search(r"[?&]ID=(\d+)", href, flags=re.IGNORECASE)
        if m:
            return m.group(1)

    txt = detail_soup.get_text(" ", strip=True)
    m = re.search(r"[?&]ID=(\d+)", txt, flags=re.IGNORECASE)
    if m:
        return m.group(1)

    if fallback_urls:
        for u in fallback_urls:
            if not u:
                continue
            m = re.search(r"[?&]ID=(\d+)", u, flags=re.IGNORECASE)
            if m:
                return m.group(1)

    return ""

def _build_schools_url(entity_id: str) -> str:
    eid = _digits_only(entity_id)
    if not eid:
        return ""
    return urljoin(EDNA_BASE, f"/Screens/Details/wfSchools.aspx?ID={eid}")

def prepopulate_edna_from_districts(xlsx_path: str, sheets: list[str], delay_sec: float = 0.6):
    """
    Crawl district → schools listings and populate edna_output.csv.
    Now captures Grades by visiting each school detail page (direct-link rows).
    """
    districts = _collect_unique_district_names_from_workbook(xlsx_path, sheets)
    if not districts:
        print("[district-prep] No districts found in workbook; skipping prepopulation.")
        return

    session = _make_session()
    print(f"[district-prep] Starting EDNA prepopulation for {len(districts)} districts")

    for lea_name in tqdm(districts, desc="Pre-crawling districts"):
        if not lea_name:
            continue

        try:
            candidates, search_url = _search_currentname(session, lea_name)
        except Exception as e:
            print(f"[district-prep] search failed for '{lea_name}': {e}")
            traceback.print_exc()
            continue

        if not candidates:
            print(f"[district-prep] no candidates for '{lea_name}'")
            continue

        district_rows = [(nm, br, href) for (nm, br, href) in candidates if _normalize_space(br) == "0000"]
        if not district_rows:
            print(f"[district-prep] no '0000' district rows for '{lea_name}'")
            continue

        for (inst_name, _branch, href) in district_rows:
            try:
                time.sleep(delay_sec)
                detail_soup, detail_url = _fetch_detail_soup(session, search_url, href)
            except Exception as e:
                print(f"[district-prep] detail fetch failed for '{inst_name}': {e}")
                traceback.print_exc()
                continue
            if not detail_soup:
                print(f"[district-prep] empty detail soup for '{inst_name}'")
                continue

            district7 = _extract_district_nces_from_details(detail_soup)
            if len(_digits_only(district7)) != 7:
                print(f"[district-prep] '{inst_name}' missing 7-digit District NCES (saw '{district7}')")
                continue

            tab = _open_schools_branches_direct(session, detail_soup, referer_url=(detail_url or search_url))
            if not tab:
                print(f"[district-prep] could not open Schools/Branches for '{inst_name}'")
                continue
            schools_soup, referer_url = tab

            page_num = 1
            while True:
                rows = _schools_table_parse(schools_soup) or []   
                if not rows:
                    print(f"[schools] zero-rows url={referer_url} inst=\"{inst_name}\" district7={_digits_only(district7)}")
                
                if not rows and page_num == 1:
                    try:
                        txt = str(schools_soup)
                        print(f"[debug:schools_tab] head:\n{txt[:1500]}\n--- [truncated len={len(txt)}] ---", flush=True)
                    except Exception:
                        pass

                for row in rows:
                    if _normalize_space(row.get("branch", "")) == "0000":
                        continue
                    _append_school_row_from_listing(inst_name, _digits_only(district7), row, session, referer_url)

                try:
                    nxt = _paginate_next(session, referer_url, schools_soup)
                except Exception as e:
                    print(f"[district-prep] pagination error on '{inst_name}' page {page_num}: {e}")
                    traceback.print_exc()
                    nxt = None

                if not nxt:
                    break
                schools_soup, referer_url = nxt
                page_num += 1
                time.sleep(delay_sec)

# ==============================
# Main
# ==============================
def main():
    try:
        _ensure_output_csv_exists()

        wb = load_workbook(CMP_FILENAME)

        # Prepopulate cache (now includes Grades)
        try:
            prepopulate_edna_from_districts(CMP_FILENAME, CMP_SHEETS, delay_sec=0.6)
        except Exception as e:
            print(f"[district-prep] {e}")
            traceback.print_exc()

        # RELOAD LOOKUP *after* prepopulation
        lookup = pd.read_csv(EDNA_CACHE_CSV, dtype=str).fillna("")
        ensure_headers(
            lookup,
            ["School Name", "District Name", "NCES 12-digit (District+Branch)"],
            "output.csv"
        )
        # Ensure 'Grades' column exists for downstream mapping
        if "Grades" not in lookup.columns:
            lookup["Grades"] = ""
            
        rows12, rowsDist7, rowsGrades = [], [], []
        for _, r in lookup.iterrows():
            k = _pair_key(r.get("School Name",""), r.get("District Name",""))
            code12 = _digits_only(r.get("NCES 12-digit (District+Branch)",""))
            dist7_csv = _digits_only(r.get("District NCES",""))
            grades = r.get("Grades", "") or ""
            if code12:
                rows12.append((k, code12))
                rowsDist7.append((k, dist7_csv if len(dist7_csv)==7 else _derive_district7_from_12(code12)))
                rowsGrades.append((k, grades))

        pair_to_nces12 = dict(rows12)
        pair_to_dist7  = dict(rowsDist7)
        pair_to_grades = dict(rowsGrades)

        # 3) Process each sheet
        for sheet in CMP_SHEETS:
            print(f"Processing sheet: {sheet}")

            df = pd.read_excel(CMP_FILENAME, sheet_name=sheet, dtype=str).fillna("")
            df.rename(columns=lambda c: c.strip() if isinstance(c, str) else c, inplace=True)

            ensure_headers(df, ["School Name", "District Name"], f"{sheet}")
            if "School Number (NCES)" not in df.columns:
                df["School Number (NCES)"] = ""
            if "District Number (NCES)" not in df.columns:
                df["District Number (NCES)"] = ""
            if "Grade Band" not in df.columns:
                df["Grade Band"] = ""
                
            if sheet == "School Pop. Data":
                if "Lowest Grade Level Served" not in df.columns:
                    df["Lowest Grade Level Served"] = ""
                if "Highest Grade Level Served" not in df.columns:
                    df["Highest Grade Level Served"] = ""                        

            for idx, row in tqdm(df.iterrows(), total=len(df), desc=f"Matching names in {sheet}"):
                school = row.get("School Name", "")
                district = row.get("District Name", "")
                if not norm(school) or not norm(district):
                    print(f"[WARN] Missing School or District at row {idx+2} in '{sheet}'")
                    continue

                key = _pair_key(school, district)

                # ---- Exact (local CSV) match on 12-digit
                nces12 = pair_to_nces12.get(key, "")
                if nces12:
                    df.at[idx, "School Number (NCES)"]   = nces12
                    dist7 = pair_to_dist7.get(key, "") or _derive_district7_from_12(nces12)
                    df.at[idx, "District Number (NCES)"] = dist7
                    
                    # Grade band from cache if available
                    gb = pair_to_grades.get(key, "")
                    if gb:
                        df.at[idx, "Grade Band"] = gb

                    # If we're on the School Pop. Data sheet, derive min/max
                    if sheet == "School Pop. Data":
                        tokens = _parse_grade_tokens(df.at[idx, "Grade Band"])
                        lo, hi = _lowest_highest_from_tokens(tokens)
                        if lo:
                            df.at[idx, "Lowest Grade Level Served"] = lo
                        if hi:
                            df.at[idx, "Highest Grade Level Served"] = hi
                            
                    continue

                print(f"[INFO] No exact match for row {idx+2} in '{sheet}': "
                      f"School='{norm(school)}', District='{norm(district)}'")

                # ---- Online lookup by name (returns grades now)
                web_row = edna_lookup_by_name(norm(school), norm(district), delay_sec=0.6)
                if web_row and (web_row.get("NCES 12-digit (District+Branch)") or web_row.get("NCES Code")):
                    code12 = _digits_only(web_row.get("NCES 12-digit (District+Branch)", "")) or _digits_only(web_row.get("NCES Code", ""))
                    dist7  = _digits_only(web_row.get("District NCES", "")) or _derive_district7_from_12(code12)
                    grades = web_row.get("Grades", "") or ""
                    df.at[idx, "School Number (NCES)"]   = code12
                    df.at[idx, "District Number (NCES)"] = dist7

                    if grades:
                        df.at[idx, "Grade Band"] = grades
                        
                    # If we're on the School Pop. Data sheet, derive min/max
                    if sheet == "School Pop. Data":
                        tokens = _parse_grade_tokens(df.at[idx, "Grade Band"])
                        lo, hi = _lowest_highest_from_tokens(tokens)
                        if lo:
                            df.at[idx, "Lowest Grade Level Served"] = lo
                        if hi:
                            df.at[idx, "Highest Grade Level Served"] = hi                         

                    try:
                        _append_if_new(web_row)
                    except Exception as e:
                        print(f"[online-append/name] {e}")
                        traceback.print_exc() 
                            
                    continue

                # ---- Fallback by LOCATION_ID (returns grades now)
                loc_candidates = [
                    row.get("LOCATION_ID", ""),
                    row.get("Location ID", ""),
                    row.get("EDNA Location ID", ""),
                    row.get("SchoolBranch", ""),
                    row.get("School Branch", ""),
                    row.get("School/Branch", ""),
                ]
                loc_raw = next((str(x) for x in loc_candidates if norm(x)), "")
                loc_digits = _digits_only(loc_raw)
                if loc_digits:
                    print(f"[INFO] Attempting location_id lookup for row {idx+2} "
                          f"in '{sheet}' using SchoolBranch={loc_digits.zfill(4)}")
                    web_row_loc = edna_lookup_by_location_id(loc_digits, delay_sec=0.6)
                    if web_row_loc and (web_row_loc.get("NCES 12-digit (District+Branch)") or web_row_loc.get("NCES Code")):
                        code12 = _digits_only(web_row_loc.get("NCES 12-digit (District+Branch)", "")) or _digits_only(web_row_loc.get("NCES Code", ""))
                        dist7  = _digits_only(web_row_loc.get("District NCES", "")) or _derive_district7_from_12(code12)
                        grades = web_row_loc.get("Grades", "") or ""
                        df.at[idx, "School Number (NCES)"]   = code12
                        df.at[idx, "District Number (NCES)"] = dist7
                        
                        if grades:
                            df.at[idx, "Grade Band"] = grades

                        # If we're on the School Pop. Data sheet, derive min/max
                        if sheet == "School Pop. Data":
                            tokens = _parse_grade_tokens(df.at[idx, "Grade Band"])
                            lo, hi = _lowest_highest_from_tokens(tokens)
                            if lo:
                                df.at[idx, "Lowest Grade Level Served"] = lo
                            if hi:
                                df.at[idx, "Highest Grade Level Served"] = hi                            
                            
                        try:
                            _append_if_new(web_row_loc)
                        except Exception as e:
                            print(f"[online-append/location_id] {e}")
                            traceback.print_exc()
                        continue

                # ---- Optional fuzzy fallback
                if ENABLE_FUZZY_MATCH and pair_to_nces12:
                    csv_keys = list(pair_to_nces12.keys())
                    best_match, score = _extract_one(key, csv_keys)
                    if best_match and score >= FUZZY_THRESHOLD:
                        code12 = pair_to_nces12.get(best_match, "")
                        df.at[idx, "School Number (NCES)"]   = code12
                        df.at[idx, "District Number (NCES)"] = pair_to_dist7.get(best_match, "") or _derive_district7_from_12(code12)
                        gb = pair_to_grades.get(best_match, "")
                        
                        if gb:
                            df.at[idx, "Grade Band"] = gb
                            
                        if sheet == "School Pop. Data":
                            tokens = _parse_grade_tokens(df.at[idx, "Grade Band"])
                            lo, hi = _lowest_highest_from_tokens(tokens)
                            if lo:
                                df.at[idx, "Lowest Grade Level Served"] = lo
                            if hi:
                                df.at[idx, "Highest Grade Level Served"] = hi                            
                            
                        print(f"[FUZZY] Using fuzzy match (score {score:.1f}) → {best_match}")
                        continue

                print(f"[WARN] No match found for row {idx+2} in '{sheet}': "
                      f"School='{norm(school)}', District='{norm(district)}'")

            # 5) Write NCES + Grade Band back to THIS sheet now
            ws = wb[sheet]
            header_to_col = build_header_map(ws)

            if "School Number (NCES)" not in header_to_col:
                col_idx = len(header_to_col) + 1
                ws.cell(row=1, column=col_idx, value="School Number (NCES)")
                header_to_col["School Number (NCES)"] = col_idx
            if "District Number (NCES)" not in header_to_col:
                col_idx = len(header_to_col) + 1
                ws.cell(row=1, column=col_idx, value="District Number (NCES)")
                header_to_col["District Number (NCES)"] = col_idx
            if "Grade Band" not in header_to_col:
                col_idx = len(header_to_col) + 1
                ws.cell(row=1, column=col_idx, value="Grade Band")
                header_to_col["Grade Band"] = col_idx

            # Only for School Pop. Data: ensure Lowest/Highest columns in the sheet
            if sheet == "School Pop. Data":
                if "Lowest Grade Level Served" not in header_to_col:
                    col_idx = len(header_to_col) + 1
                    ws.cell(row=1, column=col_idx, value="Lowest Grade Level Served")
                    header_to_col["Lowest Grade Level Served"] = col_idx
                if "Highest Grade Level Served" not in header_to_col:
                    col_idx = len(header_to_col) + 1
                    ws.cell(row=1, column=col_idx, value="Highest Grade Level Served")
                    header_to_col["Highest Grade Level Served"] = col_idx

            # ---- Column indices ----
            nces_school_col   = header_to_col["School Number (NCES)"]
            nces_district_col = header_to_col["District Number (NCES)"]
            grade_band_col    = header_to_col["Grade Band"]
            low_col = header_to_col.get("Lowest Grade Level Served")
            high_col = header_to_col.get("Highest Grade Level Served")

            # ---- Force text for NCES columns ----
            for col_idx in (nces_school_col, nces_district_col):
                for r in range(2, len(df) + 2):
                    ws.cell(row=r, column=col_idx).number_format = "@"

            # ---- Write values ----
            for r in tqdm(range(len(df)), total=len(df), desc=f"Writing outputs in {sheet}", leave=False):
                excel_row = r + 2
                ws.cell(row=excel_row, column=nces_school_col,   value=df.iloc[r]["School Number (NCES)"])
                ws.cell(row=excel_row, column=nces_district_col, value=df.iloc[r]["District Number (NCES)"])
                ws.cell(row=excel_row, column=grade_band_col,    value=df.iloc[r]["Grade Band"])
                if sheet == "School Pop. Data":
                    if low_col:
                        ws.cell(row=excel_row, column=low_col,  value=df.iloc[r]["Lowest Grade Level Served"])
                    if high_col:
                        ws.cell(row=excel_row, column=high_col, value=df.iloc[r]["Highest Grade Level Served"])

        out_name = OUTPUT_FILENAME
        wb.save(out_name)
        print(f"File saved as '{out_name}'")

    except Exception as e:
        print(f"[main] {e}")
        traceback.print_exc()

if __name__ == "__main__":
    main()
