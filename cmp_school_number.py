import pandas as pd
import yaml # pip install pyyaml
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor
from rapidfuzz import process, fuzz # pip install rapidfuzz
import re

# Load configuration from YAML
with open('config.yaml', 'r') as f:
    config = yaml.safe_load(f)
    
cmp_file_path_prefix = config['cmp_file_path_prefix']
elsi_school_file_name = f"{cmp_file_path_prefix}/{config['elsi_school_file_name']}"
elsi_district_file_name = f"{cmp_file_path_prefix}/{config['elsi_district_file_name']}"
cmp_output_file = f"{cmp_file_path_prefix}/{config['cmp_output_file_name']}"

output_school_name_col = "School Name"
output_district_name_col = "District Name"
output_school_id_col = "School Number (NCES)"
output_district_id_col = "District Number (NCES)"

elsi_school_col = "School Name"
elsi_district_col = "Agency Name"
elsi_school_id_col = "School ID (12-digit) - NCES Assigned [Public School] Latest available year"
elsi_district_id_col = "Agency ID - NCES Assigned [District] Latest available year"

def normalize_name(name):
    if not name:
        return ''
    name = name.lower()
    name = re.sub(r'[^\w\s-]', '', name)  # remove punctuation
    name = re.sub(r'\s+', ' ', name)     # collapse whitespace
    return name.strip()

def get_id_from_name(name, source_df, name_col, id_col, threshold=70):
    if not name:
        return None

    # Normalize the query and choices
    normalized_query = normalize_name(name)
    choices = source_df[name_col].astype(str)
    normalized_choices = choices.map(normalize_name).tolist()

    # Use normalized names for matching
    best_match = process.extractOne(normalized_query, normalized_choices,
                                     scorer=fuzz.token_sort_ratio,
                                     score_cutoff=threshold)

    if best_match:
        match, score, match_index = best_match
        #print(f"Matched {name} to {match}")
        return source_df.iloc[match_index][id_col]

    print(f"No match for: '{name}' (normalized: '{normalized_query}')")
    return None
    
def match_row(cells, idx, col_idx, school_df, district_df):
    school_name = cells[col_idx['school_name'] - 1].value
    district_name = cells[col_idx['district_name'] - 1].value

    school_id = get_id_from_name(school_name, school_df,
                                 elsi_school_col,
                                 elsi_school_id_col)

    district_id = get_id_from_name(district_name, district_df,
                                   elsi_district_col,
                                   elsi_district_id_col)

    return cells, col_idx['school_id'], school_id, col_idx['district_id'], district_id

def populate_ids(output_path, district_source_path, school_source_path,
                 output_name_col, id_col_to_fill,
                 match_on_district=True):
    # Load dataframes, skipping first 6 rows
    output_df = pd.read_excel(output_path)
    district_df = pd.read_excel(district_source_path, skiprows=6)
    school_df = pd.read_excel(school_source_path, skiprows=6)

    # Choose source and columns based on type of ID
    if match_on_district:
        source_df = district_df
        name_col = elsi_district_col
        id_col = elsi_district_id_col
    else:
        source_df = school_df
        name_col = elsi_school_col
        id_col = elsi_school_id_col

    # Fill in the ID column
    for idx, row in output_df.iterrows():
        name_to_match = row[output_name_col]
        matched_id = get_id_from_name(name_to_match, source_df, name_col, id_col)
        output_df.at[idx, id_col_to_fill] = matched_id

    return output_df

# Load source Excel data (data starts on row 7 in the source files)
district_df = pd.read_excel(elsi_district_file_name, skiprows=6)
school_df = pd.read_excel(elsi_school_file_name, skiprows=6)

# Load output file with openpyxl
wb = load_workbook(cmp_output_file)

for output_sheet_name in ["School Pop. Data", "School CS Data"]:
    ws = wb[output_sheet_name]

    # Load header row from output file to find column indices
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    # Get column indices (1-based for openpyxl)
    col_idx = {
        'school_name': header.index(output_school_name_col) + 1,
        'district_name': header.index(output_district_name_col) + 1,
        'school_id': header.index(output_school_id_col) + 1,
        'district_id': header.index(output_district_id_col) + 1,
    }

    # Phase 1: Collect all data rows (value-only)
    rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row)) # Start from row 2 (since row 1 is the header)

    # Phase 2: Run matching in parallel
    results = []
    with ThreadPoolExecutor(max_workers=8) as executor:
        futures = [
            executor.submit(match_row, row, idx + 2, col_idx, school_df, district_df)
            for idx, row in enumerate(rows)
        ]
        for future in futures:
            results.append(future.result())

    # Phase 3: Write results to Excel (sequential)
    for cells, school_col, school_id, district_col, district_id in results:
        #print("Writing", school_col, school_id, district_col, district_id)
        if school_id is not None:
            cells[school_col - 1].value = school_id
        if district_id is not None:
            cells[district_col - 1].value = district_id

    # Save in-place 
    wb.save(cmp_output_file)
