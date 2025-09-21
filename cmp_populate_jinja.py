import sqlite3
from jinja2 import Template
import yaml  # pip install pyyaml
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os
import traceback
from tqdm import tqdm

# =========================
# Config load
# =========================
try:
    with open('config.yaml', 'r') as f:
        config = yaml.safe_load(f)
except Exception as e:
    print(f"[config] Failed to load config.yaml: {e}")
    traceback.print_exc()
    raise

# Years / files
school_year_dash  = config['years']
school_year_splat = [yr.replace('-', '_') for yr in school_year_dash]
file_path_prefix  = config['file_path_prefix']
cmp_output_file   = f"{file_path_prefix}/{config['cmp_output_file_name']}"
db_file           = f"{file_path_prefix}/{config['db_file_name']}"

# -------- Output column names (what you want in the Excel workbook) --------
OUT_SCHOOL_NAME  = config.get('output_school_name_col',  'School Name')
OUT_DISTRICT_NAME= config.get('output_district_name_col','District Name')
OUT_SCHOOL_NCES  = config.get('output_school_id_col',    'School Number (NCES)')
OUT_DISTRICT_NCES= config.get('output_district_id_col',  'District Number (NCES)')
OUT_LOW_GRADE    = config.get('output_low_grade_band_col','Lowest Grade Level Served')
OUT_HIGH_GRADE   = config.get('output_high_grade_band_col','Highest Grade Level Served')

# -------- ELSI *source* column names (as they appear in your data/sql joins) --------
ELSI_SCHOOL_ID   = config.get('elsi_school_id_col')       
ELSI_DISTRICT_ID = config.get('elsi_district_id_col')      
ELSI_LOW_GRADE   = config.get('elsi_low_grade_band',  OUT_LOW_GRADE)
ELSI_HIGH_GRADE  = config.get('elsi_high_grade_band', OUT_HIGH_GRADE)

# =========================
# Schemas (built from configured output names)
# =========================
SCHEMA_CS = [
    "School Year",
    OUT_SCHOOL_NCES, f"{OUT_SCHOOL_NAME} (State)" if OUT_SCHOOL_NAME else "School Number (State)",
    OUT_DISTRICT_NCES, f"{OUT_DISTRICT_NAME} (State)" if OUT_DISTRICT_NAME else "District Number (State)",
    OUT_LOW_GRADE, OUT_HIGH_GRADE,
    "Category",
    "Basic_Courses", "Basic_Total",
    "Adv_Courses", "Adv_Total"
]

SCHEMA_POP = [
    "School Year",
    OUT_SCHOOL_NCES, f"{OUT_SCHOOL_NAME} (State)" if OUT_SCHOOL_NAME else "School Number (State)",
    OUT_DISTRICT_NCES, f"{OUT_DISTRICT_NAME} (State)" if OUT_DISTRICT_NAME else "District Number (State)",
    OUT_LOW_GRADE, OUT_HIGH_GRADE,
    "Girls", "Boys", "Gender X", "Total"
]

# =========================
# Utilities
# =========================
def _canon(name: str) -> str:
    s = str(name) if name is not None else ""
    s = s.replace("\xa0", " ")
    s = " ".join(s.split())
    return s.strip().lower()

def _ensure_parent_dir(path: str):
    parent = os.path.dirname(os.path.abspath(path))
    if parent and not os.path.exists(parent):
        os.makedirs(parent, exist_ok=True)

def _first_present(df: pd.DataFrame, candidates: list[str]) -> str | None:
    for c in candidates:
        if c and c in df.columns:
            return c
    return None

def _parse_grade_band_from_grades_column(df: pd.DataFrame, grades_col: str = 'Grades') -> tuple[pd.Series, pd.Series]:
    """
    Derive Lowest/Highest grade from a comma-separated 'Grades' column.
    e.g., '1,2,3,4,5' -> (1, 5). Non-numeric tokens ignored. Empty -> <NA>.
    """
    lowest = pd.Series(pd.NA, index=df.index, dtype='Int64')
    highest = pd.Series(pd.NA, index=df.index, dtype='Int64')

    if grades_col not in df.columns:
        return lowest, highest

    def parse_one(val):
        if pd.isna(val):
            return (pd.NA, pd.NA)
        if isinstance(val, str):
            tokens = [t.strip() for t in val.split(',')]
        elif isinstance(val, (list, tuple)):
            tokens = list(val)
        else:
            tokens = [str(val)]

        nums = []
        for t in tokens:
            if isinstance(t, str) and t.isdigit():
                nums.append(int(t))
            elif isinstance(t, (int, float)) and not pd.isna(t):
                try:
                    nums.append(int(t))
                except Exception:
                    pass

        if not nums:
            return (pd.NA, pd.NA)
        return (min(nums), max(nums))

    parsed = [parse_one(v) for v in df[grades_col]]
    if parsed:
        lows, highs = zip(*parsed)
        lowest = pd.Series(lows, index=df.index, dtype='Int64')
        highest = pd.Series(highs, index=df.index, dtype='Int64')
    return lowest, highest

# =========================
# Workbook init
# =========================
def init_workbook(filepath: str,
                  sheets_to_headers: dict[str, list[str]],
                  header_row: int = 1,
                  overwrite: bool = True):
    """
    Create (or overwrite) an .xlsx workbook with specified sheets and headers.
    """
    _ensure_parent_dir(filepath)
    if overwrite and os.path.exists(filepath):
        os.remove(filepath)

    wb = Workbook()
    first = True
    for sheet_name, headers in sheets_to_headers.items():
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name)

        # Write headers
        for idx, h in enumerate(headers, start=1):
            ws.cell(row=header_row, column=idx, value=h)

        # Freeze panes below header
        ws.freeze_panes = f"A{header_row+1}"

        # AutoFilter
        last_col = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A{header_row}:{last_col}{header_row}"

    wb.save(filepath)

# =========================
# Normalizers (config-aware)
# =========================
def normalize_school_pop_df(df: pd.DataFrame, school_year_value: str) -> pd.DataFrame:
    try:
        df = df.copy()
        df['School Year'] = school_year_value

        # ---- Mirror NCES -> "State" ID placeholders (writer expects columns present) ----
        # Output ID columns
        if OUT_SCHOOL_NCES in df.columns:
            # Create a placeholder "state" column to satisfy schema; it will mirror NCES
            df[f"{OUT_SCHOOL_NAME} (State)"] = df[OUT_SCHOOL_NCES]
        else:
            if f"{OUT_SCHOOL_NAME} (State)" not in df.columns:
                df[f"{OUT_SCHOOL_NAME} (State)"] = pd.NA

        if OUT_DISTRICT_NCES in df.columns:
            df[f"{OUT_DISTRICT_NAME} (State)"] = df[OUT_DISTRICT_NCES]
        else:
            if f"{OUT_DISTRICT_NAME} (State)" not in df.columns:
                df[f"{OUT_DISTRICT_NAME} (State)"] = pd.NA

        # Robust numeric helper
        def _num(col: str) -> pd.Series:
            if col in df.columns:
                return pd.to_numeric(df[col], errors='coerce').fillna(0)
            return pd.Series(0, index=df.index, dtype='float64')

        # Compute Total if not present
        if 'Total' not in df.columns:
            total = _num('Girls') + _num('Boys') + _num('Gender X')
            df['Total'] = total.round().astype('Int64')

        # --- Populate grade bands into OUTPUT columns ---
        # Prefer explicit *ELSI source* columns, else any pre-existing OUTPUT columns, else derive from Grades.
        low_src  = _first_present(df, [ELSI_LOW_GRADE, OUT_LOW_GRADE])
        high_src = _first_present(df, [ELSI_HIGH_GRADE, OUT_HIGH_GRADE])

        if low_src and low_src != OUT_LOW_GRADE:
            df[OUT_LOW_GRADE] = df[low_src]
        elif OUT_LOW_GRADE not in df.columns:
            df[OUT_LOW_GRADE] = pd.NA

        if high_src and high_src != OUT_HIGH_GRADE:
            df[OUT_HIGH_GRADE] = df[high_src]
        elif OUT_HIGH_GRADE not in df.columns:
            df[OUT_HIGH_GRADE] = pd.NA

        # If still blank, try deriving from Grades
        need_low = df[OUT_LOW_GRADE].isna().all() if OUT_LOW_GRADE in df.columns else True
        need_high = df[OUT_HIGH_GRADE].isna().all() if OUT_HIGH_GRADE in df.columns else True
        if need_low or need_high:
            grades_col = 'Grades' if 'Grades' in df.columns else ('grades' if 'grades' in df.columns else None)
            if grades_col is not None:
                low_s, high_s = _parse_grade_band_from_grades_column(df, grades_col=grades_col)
                if OUT_LOW_GRADE in df.columns:
                    df[OUT_LOW_GRADE] = df[OUT_LOW_GRADE].fillna(low_s)
                else:
                    df[OUT_LOW_GRADE] = low_s
                if OUT_HIGH_GRADE in df.columns:
                    df[OUT_HIGH_GRADE] = df[OUT_HIGH_GRADE].fillna(high_s)
                else:
                    df[OUT_HIGH_GRADE] = high_s

        return df
    except Exception as e:
        print(f"[normalize_school_pop_df] {e}")
        traceback.print_exc()
        return df

def normalize_school_cs_df(df: pd.DataFrame, school_year_value: str) -> pd.DataFrame:
    """
    Ensures School CS rows have standardized IDs, filters out zero-only course rows,
    and populates OUTPUT grade-band columns using config-aware sources.
    """
    try:
        df = df.copy()
        df['School Year'] = school_year_value

        # Ensure OUTPUT ID columns exist; mirror to the "(State)" placeholders
        for col in [OUT_SCHOOL_NCES, OUT_DISTRICT_NCES, f"{OUT_SCHOOL_NAME} (State)", f"{OUT_DISTRICT_NAME} (State)"]:
            if col not in df.columns:
                df[col] = pd.NA
        if OUT_SCHOOL_NCES in df.columns:
            df[f"{OUT_SCHOOL_NAME} (State)"] = df[OUT_SCHOOL_NCES]
        if OUT_DISTRICT_NCES in df.columns:
            df[f"{OUT_DISTRICT_NAME} (State)"] = df[OUT_DISTRICT_NCES]

        # Category presence
        if 'Category' not in df.columns:
            df['Category'] = pd.NA

        # --- Populate grade bands into OUTPUT columns (prefer ELSI source names) ---
        low_src  = _first_present(df, [ELSI_LOW_GRADE, OUT_LOW_GRADE])
        high_src = _first_present(df, [ELSI_HIGH_GRADE, OUT_HIGH_GRADE])

        if low_src and low_src != OUT_LOW_GRADE:
            df[OUT_LOW_GRADE] = df[low_src]
        elif OUT_LOW_GRADE not in df.columns:
            df[OUT_LOW_GRADE] = pd.NA

        if high_src and high_src != OUT_HIGH_GRADE:
            df[OUT_HIGH_GRADE] = df[high_src]
        elif OUT_HIGH_GRADE not in df.columns:
            df[OUT_HIGH_GRADE] = pd.NA

        # Fallback: derive from Grades
        need_low = df[OUT_LOW_GRADE].isna().all() if OUT_LOW_GRADE in df.columns else True
        need_high = df[OUT_HIGH_GRADE].isna().all() if OUT_HIGH_GRADE in df.columns else True
        if need_low or need_high:
            grades_col = 'Grades' if 'Grades' in df.columns else ('grades' if 'grades' in df.columns else None)
            if grades_col is not None:
                low_s, high_s = _parse_grade_band_from_grades_column(df, grades_col=grades_col)
                if OUT_LOW_GRADE in df.columns:
                    df[OUT_LOW_GRADE] = df[OUT_LOW_GRADE].fillna(low_s)
                else:
                    df[OUT_LOW_GRADE] = low_s
                if OUT_HIGH_GRADE in df.columns:
                    df[OUT_HIGH_GRADE] = df[OUT_HIGH_GRADE].fillna(high_s)
                else:
                    df[OUT_HIGH_GRADE] = high_s

        # Filter out rows where all CS counts are zero
        def _num(col: str) -> pd.Series:
            if col in df.columns:
                return pd.to_numeric(df[col], errors='coerce').fillna(0)
            return pd.Series(0, index=df.index, dtype='float64')

        basic_courses = _num('Basic_Courses')
        basic_total   = _num('Basic_Total')
        adv_courses   = _num('Adv_Courses')
        adv_total     = _num('Adv_Total')

        keep = ~((basic_courses == 0) & (basic_total == 0) &
                 (adv_courses == 0)   & (adv_total == 0))
        df = df[keep].reset_index(drop=True)

        return df
    except Exception as e:
        print(f"[normalize_school_cs_df] {e}")
        traceback.print_exc()
        return df

# =========================
# Excel writer
# =========================
def write_dataframes_to_excel(filepath: str,
                              sheet_name: str,
                              dataframes: list[pd.DataFrame],
                              header_row: int = 1,
                              start_row: int = 2,
                              declared_headers: list[str] | None = None):
    """
    Write multiple DataFrames into an existing sheet, clearing previous data rows.
    """
    def escape_for_excel(val):
        if pd.isnull(val):
            return ''
        s = str(val)
        if '-' in s or '/' in s:
            return f'="{s}"'
        return s

    try:
        wb = load_workbook(filepath)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in '{filepath}'.")

        ws = wb[sheet_name]

        # Build header maps (original + canonical)
        header_cells = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=False))[0]
        header_index = {}
        header_index_canon = {}
        for idx, cell in enumerate(header_cells, start=1):
            if cell.value is None:
                continue
            original = str(cell.value)
            canon = _canon(original)
            header_index[original] = idx
            if canon not in header_index_canon:
                header_index_canon[canon] = idx

        # Determine extras
        extras = []
        seen_canon = set(header_index_canon.keys())
        for df in (dataframes or []):
            for col in df.columns:
                c = _canon(col)
                if c not in seen_canon:
                    extras.append(col)
                    seen_canon.add(c)

        # Append extras
        if extras:
            start_col = ws.max_column + 1
            for offset, col_name in enumerate(extras):
                ws.cell(row=header_row, column=start_col + offset, value=col_name)

            header_cells = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=False))[0]
            header_index.clear()
            header_index_canon.clear()
            for idx, cell in enumerate(header_cells, start=1):
                if cell.value is None:
                    continue
                original = str(cell.value)
                canon = _canon(original)
                header_index[original] = idx
                if canon not in header_index_canon:
                    header_index_canon[canon] = idx

        # Clear old data rows
        if ws.max_row > (start_row - 1):
            ws.delete_rows(start_row, ws.max_row - (start_row - 1))

        # Prepare write order
        if declared_headers is None:
            core_order = [str(c.value) for c in header_cells if c.value is not None]
        else:
            core_order = [h for h in declared_headers if _canon(h) in header_index_canon]

        extra_order = [e for e in extras if _canon(e) in header_index_canon]
        write_order = core_order + [e for e in extra_order if e not in core_order]

        # Write rows from all DataFrames
        r = start_row
        total_rows = sum(len(df) for df in (dataframes or []))
        row_progress = tqdm(total=total_rows, desc=f"Writing rows to '{sheet_name}'", unit="row")

        for df in tqdm((dataframes or []), desc=f"DataFrames -> '{sheet_name}'", unit="df"):
            df_cols = list(df.columns)
            writable_cols = [col for col in write_order if _canon(col) in {_canon(c) for c in df_cols}]
            canon_map = {_canon(c): c for c in df_cols}

            for _, row in df.iterrows():
                for col in writable_cols:
                    col_idx = header_index_canon.get(_canon(col))
                    if col_idx is None:
                        continue
                    src_col = canon_map.get(_canon(col))
                    val = row.get(src_col, '')
                    val = escape_for_excel(val)
                    cell = ws.cell(row=r, column=col_idx, value=val)
                    cell.number_format = '@'
                r += 1
                row_progress.update(1)

        row_progress.close()
        wb.save(filepath)
        print(f"Successfully wrote sheet '{sheet_name}' in '{filepath}'.")

    except Exception as e:
        print(f"[write_dataframes_to_excel] {e}")
        traceback.print_exc()
        raise

# =========================
# Main
# =========================
def main():
    cs_dfs = []
    pop_dfs = []

    try:
        conn = sqlite3.connect(db_file)
    except Exception as e:
        print(f"[sqlite3.connect] {e}")
        traceback.print_exc()
        raise

    try:
        scripts = ['school_cs_jinja.sql', 'school_pop_jinja.sql']

        for script in tqdm(scripts, desc="SQL scripts", unit="file"):
            try:
                with open(script) as file:
                    template = Template(file.read())
            except Exception as e:
                print(f"[read {script}] {e}")
                traceback.print_exc()
                raise

            for i in tqdm(range(len(school_year_dash)), desc=f"Years for {os.path.basename(script)}", unit="year"):
                dash = school_year_dash[i]
                splat = school_year_splat[i]

                try:
                    rendered_sql = template.render(
                        school_year_dash=dash,
                        school_year_splat=splat,
                        high_school_only=True  # CMP is grades 9-12 only
                    )

                    cursor = conn.cursor()
                    cursor.execute(rendered_sql)

                    results = cursor.fetchall()
                    column_names = [desc[0] for desc in cursor.description]
                    df = pd.DataFrame(results, columns=column_names)

                    if script == 'school_cs_jinja.sql':
                        df = normalize_school_cs_df(df, school_year_value=dash)
                        cs_dfs.append(df)
                    elif script == 'school_pop_jinja.sql':
                        df = normalize_school_pop_df(df, school_year_value=dash)
                        pop_dfs.append(df)
                    else:
                        print("[main] Warning: unmatched script while appending DataFrame.")

                except Exception as e:
                    print(f"[query {script} for {dash}] {e}")
                    traceback.print_exc()
                    # Continue to next year rather than abort entire run
                    continue

    finally:
        try:
            conn.close()
        except Exception as e:
            print(f"[sqlite close] {e}")
            traceback.print_exc()

    # Initialize workbook with configured headers
    try:
        init_workbook(
            cmp_output_file,
            sheets_to_headers={
                "School CS Data": SCHEMA_CS,
                "School Pop. Data": SCHEMA_POP
            },
            header_row=1,
            overwrite=True  # start clean each run
        )
    except Exception as e:
        print(f"[init_workbook] {e}")
        traceback.print_exc()
        raise

    # Write CS and Population sheets with progress bars
    try:
        write_dataframes_to_excel(
            filepath=cmp_output_file,
            sheet_name="School CS Data",
            dataframes=cs_dfs,
            header_row=1,
            start_row=2,
            declared_headers=SCHEMA_CS
        )
    except Exception as e:
        print(f"[write CS] {e}")
        traceback.print_exc()
        raise

    try:
        write_dataframes_to_excel(
            filepath=cmp_output_file,
            sheet_name="School Pop. Data",
            dataframes=pop_dfs,
            header_row=1,
            start_row=2,
            declared_headers=SCHEMA_POP
        )
    except Exception as e:
        print(f"[write Pop] {e}")
        traceback.print_exc()
        raise

    print("Done.")

if __name__ == "__main__":
    main()
