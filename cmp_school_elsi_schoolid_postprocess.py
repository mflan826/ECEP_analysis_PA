import re
import pandas as pd
from fuzzywuzzy import process
from openpyxl import load_workbook
from tqdm import tqdm
import yaml
import re

# ----------------------------
# File & sheet parameters
# ----------------------------
def _load_config():
    try:
        with open('config.yaml', 'r') as f:
            return yaml.safe_load(f) or {}
    except Exception:
        return {}

_cfg = _load_config()

def _cfg_get(key, default):
    v = _cfg.get(key)
    return v if (v is not None and str(v).strip() != "") else default
    
# Paths / filenames
FILE_PREFIX         = _cfg_get("file_path_prefix", ".")
CMP_INPUT_FILENAME  = f"{file_path_prefix}/CMP_Data_Populated.xlsx"
CMP_OUTPUT_FILENAME = f"{file_path_prefix}/CMP_Data_Populated - Updated.xlsx"
ELSI_SCHOOL_FILE    = _cfg_get("elsi_school_file_name", "ELSI_excel_export_6387876530335112514653.xlsx")
ELSI_SHEET_NAME     = _cfg_get("elsi_school_sheet_name", "ELSI Export")

# Sheets to process
CMP_SHEETS = _cfg.get("cmp_sheets", ["School Pop. Data", "School CS Data"])

# CMP column names (outputs)
CMP_SCHOOL_NAME_COL = _cfg_get("output_school_name_col", "School Name")
CMP_SCHOOL_ID_COL   = _cfg_get("output_school_id_col", "School Number (NCES)")
CMP_GRADE_BAND_COL  = _cfg_get("output_grade_band_col", "Grade Band")
CMP_LOW_GRADE_COL   = _cfg_get("output_low_grade_band_col", "Lowest Grade Level Served")
CMP_HIGH_GRADE_COL  = _cfg_get("output_high_grade_band_col", "Highest Grade Level Served")

# ELSI column names (inputs)
ELSI_SCHOOL_NAME_COL = _cfg_get("elsi_school_col", "School Name")
ELSI_SCHOOL_ID_COL   = _cfg_get("elsi_school_id_col", "School ID (12-digit) - NCES Assigned [Public School] Latest available year")
ELSI_LOW_COL_PREF    = _cfg_get("elsi_low_grade_band", "")
ELSI_HIGH_COL_PREF   = _cfg_get("elsi_high_grade_band", "")

# Fuzzy threshold
FUZZY_THRESHOLD = 60

# ----------------------------
# Helpers for grade band
# ----------------------------
_GRADE_CANON = {
    "PK": "PK", "PREK": "PK", "PRE-K": "PK", "PRE K": "PK", "PREKINDERGARTEN": "PK", "PRE KINDERGARTEN": "PK",
    "K": "K", "KG": "K", "KN": "K", "KINDER": "K", "KINDERGARTEN": "K",
}
def _canon_grade_token(s):
    if s is None: return ""
    t = re.sub(r"[^A-Za-z0-9]", "", str(s)).upper()
    if t in _GRADE_CANON: return _GRADE_CANON[t]
    if re.fullmatch(r"\d{1,2}", t):
        v = int(t)
        return str(v) if 1 <= v <= 12 else ""
    return ""

def _compose_grade_band(low, high):
    lo = _canon_grade_token(low)
    hi = _canon_grade_token(high)
    if lo and hi: return f"{lo}–{hi}"
    return lo or hi or ""

# ----------------------------
# Load ELSI (skip first 6 rows so data starts at row 8)
# ----------------------------
elsi_df = pd.read_excel(f"{FILE_PREFIX}/{ELSI_SCHOOL_FILE}", sheet_name=ELSI_SHEET_NAME, skiprows=6, dtype=str)

# Validate minimal columns
_missing = [c for c in [ELSI_SCHOOL_NAME_COL, ELSI_SCHOOL_ID_COL] if c not in elsi_df.columns]
if _missing:
    raise ValueError(f"Expected columns missing from ELSI school export: {_missing}")

# Resolve low/high grade columns (best-effort; band will be blank if not found)
elsi_low_col  = CMP_LOW_GRADE_COL
elsi_high_col = CMP_HIGH_GRADE_COL

# Build name list for fuzzy matching
elsi_names = elsi_df[ELSI_SCHOOL_NAME_COL].fillna("").tolist()

# ----------------------------
# Open the CMP workbook
# ----------------------------
wb = load_workbook(f"{FILE_PREFIX}/{CMP_INPUT_FILENAME}")

for sheet in CMP_SHEETS:
    print(f"Processing sheet: {sheet}")
    df = pd.read_excel(f"{FILE_PREFIX}/{CMP_INPUT_FILENAME}", sheet_name=sheet, dtype=str).fillna("")
    
    # Ensure required logical columns exist
    if CMP_SCHOOL_NAME_COL not in df.columns:
        raise ValueError(f"Sheet '{sheet}' is missing required column '{CMP_SCHOOL_NAME_COL}'")

    # Ensure output columns exist
    if CMP_SCHOOL_ID_COL not in df.columns:
        df[CMP_SCHOOL_ID_COL] = ""
    if CMP_GRADE_BAND_COL not in df.columns:
        df[CMP_GRADE_BAND_COL] = ""
    if CMP_LOW_GRADE_COL not in df.columns:
        df[CMP_LOW_GRADE_COL] = ""
    if CMP_HIGH_GRADE_COL not in df.columns:
        df[CMP_HIGH_GRADE_COL] = ""

    # Fuzzy match each school name and fill NCES + Grade Band from ELSI
    for idx, row in tqdm(df.iterrows(), total=len(df), desc=f"Matching schools in {sheet}"):
        school_name = str(row.get(CMP_SCHOOL_NAME_COL, "") or "").strip()
        if not school_name:
            continue

        best = process.extractOne(school_name, elsi_names)
        if not best:
            continue

        best_name, score = best[0], best[1]
        if score < FUZZY_THRESHOLD:
            continue

        match_row = elsi_df[elsi_df[ELSI_SCHOOL_NAME_COL] == best_name]
        if match_row.empty:
            continue

        r0 = match_row.iloc[0]

        # NCES 12-digit (from config key)
        nces_id = r0.get(ELSI_SCHOOL_ID_COL, "") or ""
        df.at[idx, CMP_SCHOOL_ID_COL] = nces_id

        # Grade Band from low/high if available
        low  = (r0.get(elsi_low_col)  if elsi_low_col  else "") or ""
        high = (r0.get(elsi_high_col) if elsi_high_col else "") or ""
        band = _compose_grade_band(low, high)
        if band:
            df.at[idx, CMP_GRADE_BAND_COL] = band
        # Also store normalized low/high into the configured output columns
        if low:
            df.at[idx, CMP_LOW_GRADE_COL]  = _canon_grade_token(low)
        if high:
            df.at[idx, CMP_HIGH_GRADE_COL] = _canon_grade_token(high)

    # Ensure headers exist/align in the worksheet (add any new columns at their DataFrame positions)
    ws = wb[sheet]
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        if (cell.value or "").strip() != col_name:
            cell.value = col_name

    # Write all values back (row 2 onward)
    for col_idx, col_name in enumerate(df.columns, start=1):
        for row_idx, val in tqdm(
            enumerate(df[col_name], start=2),
            total=len(df),
            desc=f"Writing to workbook ({sheet}, col: {col_name})",
            leave=False
        ):
            ws.cell(row=row_idx, column=col_idx).value = val

# Save updated workbook
out_path = f"{FILE_PREFIX}/{CMP_OUTPUT_FILENAME}"
wb.save(out_path)
print(f"File saved as '{out_path}'")

